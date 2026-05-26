#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
決算書PDF -> 3期比較 + 企業価値算定 Excel出力ツール
Usage: python kessan_tool.py <PDF1> [PDF2 PDF3] <業種>
"""

import sys, re
from pathlib import Path
import pdfplumber
import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 色定数 ───────────────────────────────────────
C_HEADER   = "1F4E79"   # 濃紺
C_MID      = "2E75B6"   # 中青
C_TOTAL    = "C5E0B4"   # 薄緑（合計確定）
C_SUBTOTAL = "DDEBF7"   # 薄水（小計確定）
C_AUTO     = "FFF2CC"   # 黄（自動入力・要確認）
C_EMPTY    = "F2F2F2"   # 薄灰（未入力テンプレート）

# ═══════════════════════════════════════════════
# 一般的な科目テンプレート定義
# (表示ラベル, キー名, 行タイプ, 確定=True/要確認=False)
# ═══════════════════════════════════════════════
BS_ASSET_TMPL = [
    ("【流動資産】",           None,                    "header",   True),
    ("現金及び預金",           "現金及び預金",            "item",     False),
    ("受取手形",              "受取手形",               "item",     False),
    ("売掛金",               "売掛金",                 "item",     False),
    ("商品・製品",             "商品・製品",              "item",     False),
    ("仕掛品",               "仕掛品",                 "item",     False),
    ("原材料・貯蔵品",          "原材料・貯蔵品",           "item",     False),
    ("前払費用",              "前払費用",               "item",     False),
    ("未収入金",              "未収入金",               "item",     False),
    ("その他流動資産",          "その他流動資産",           "item",     False),
    ("流動資産合計",            "流動資産合計",            "subtotal", True),
    ("【固定資産】",            None,                    "header",   True),
    ("有形固定資産計",          "有形固定資産計",          "item",     False),
    ("　建物・構築物",          "建物・構築物",            "item",     False),
    ("　機械及び装置",          "機械及び装置",            "item",     False),
    ("　車両運搬具",            "　車両運搬具",            "item",     False),
    ("　工具器具備品",          "工具器具備品",            "item",     False),
    ("　土地",               "土地",                  "item",     False),
    ("　減価償却累計額",         "　減価償却累計額",         "item",     False),
    ("無形固定資産合計",         "無形固定資産合計",         "item",     False),
    ("投資その他の資産合計",      "投資その他資産合計",       "item",     False),
    ("　敷金・保証金",          "敷金・保証金",            "item",     False),
    ("　繰延税金資産",          "繰延税金資産",            "item",     False),
    ("　その他投資資産",         "その他投資資産",          "item",     False),
    ("固定資産合計",            "固定資産合計",            "subtotal", True),
    ("資産合計",              "資産合計",               "total",    True),
]

BS_LIAB_TMPL = [
    ("【流動負債】",            None,                    "header",   True),
    ("支払手形",              "支払手形",               "item",     False),
    ("買掛金",               "買掛金",                 "item",     False),
    ("短期借入金",             "短期借入金",              "item",     False),
    ("1年内返済長期借入金",      "1年内返済長期借入金",      "item",     False),
    ("未払金",               "未払金",                 "item",     False),
    ("未払費用",              "未払費用",               "item",     False),
    ("未払法人税等",            "未払法人税等",            "item",     False),
    ("前受金",               "前受金",                 "item",     False),
    ("預り金",               "預り金",                 "item",     False),
    ("賞与引当金",             "賞与引当金",              "item",     False),
    ("その他流動負債",          "その他流動負債",           "item",     False),
    ("流動負債合計",            "流動負債合計",            "subtotal", True),
    ("【固定負債】",            None,                    "header",   True),
    ("長期借入金",             "長期借入金",              "item",     False),
    ("退職給付引当金",          "退職給付引当金",           "item",     False),
    ("その他固定負債",          "その他固定負債",           "item",     False),
    ("固定負債合計",            "固定負債合計",            "subtotal", True),
    ("負債合計",              "負債合計",               "total",    True),
    ("【純資産の部】",          None,                    "header",   True),
    ("資本金",               "資本金",                 "item",     False),
    ("資本剰余金",             "資本剰余金",              "item",     False),
    ("利益準備金",             "利益準備金",              "item",     False),
    ("繰越利益剰余金",          "繰越利益剰余金",           "item",     False),
    ("その他利益剰余金",         "その他利益剰余金",         "item",     False),
    ("利益剰余金合計",          "利益剰余金合計",           "subtotal", True),
    ("自己株式",              "自己株式",               "item",     False),
    ("株主資本合計",            "株主資本合計",            "subtotal", True),
    ("その他包括利益累計額合計",   "その他包括利益累計額合計",  "item",     False),
    ("純資産合計",             "純資産合計",              "total",    True),
    ("負債・純資産合計",         "負債・純資産合計",         "total",    True),
]

PL_TMPL = [
    ("売上高",               "売上高",                 "total",    True),
    ("【売上原価】",           None,                    "header",   True),
    ("期首棚卸高",             "期首棚卸高",              "item",     False),
    ("当期仕入高",             "当期仕入高",              "item",     False),
    ("当期製品製造原価",         "当期製品製造原価",         "item",     False),
    ("期末棚卸高",             "期末棚卸高",              "item",     False),
    ("売上原価合計",            "売上原価合計",            "subtotal", True),
    ("売上総利益（損失）",        "売上総利益（損失）",        "total",    True),
    ("販売費及び一般管理費合計",   "販売費及び一般管理費合計",   "subtotal", True),
    ("営業利益（損失）",         "営業利益（損失）",         "total",    True),
    ("【営業外収益】",          None,                    "header",   True),
    ("受取利息・配当金",         "受取利息",               "item",     False),
    ("雑収入",               "雑収入",                 "item",     False),
    ("その他営業外収益",         "その他営業外収益",         "item",     False),
    ("営業外収益合計",          "営業外収益合計",           "subtotal", True),
    ("【営業外費用】",          None,                    "header",   True),
    ("支払利息・割引料",         "営業外費用合計",           "item",     False),
    ("為替差損",              "為替差損",               "item",     False),
    ("その他営業外費用",         "その他営業外費用",         "item",     False),
    ("営業外費用合計",          "営業外費用合計",           "subtotal", True),
    ("経常利益（損失）",         "経常利益（損失）",         "total",    True),
    ("【特別利益】",            None,                    "header",   True),
    ("固定資産売却益",          "固定資産売却益",           "item",     False),
    ("その他特別利益",          "その他特別利益",           "item",     False),
    ("特別利益合計",            "特別利益合計",            "subtotal", True),
    ("【特別損失】",            None,                    "header",   True),
    ("固定資産売却損・廃棄損",    "固定資産売却損",          "item",     False),
    ("減損損失",              "減損損失",               "item",     False),
    ("その他特別損失",          "その他特別損失",           "item",     False),
    ("特別損失合計",            "特別損失合計",            "subtotal", True),
    ("税引前当期純利益（損失）",   "税引前当期純利益（損失）",   "total",    True),
    ("法人税等",              "法人税等",               "item",     False),
    ("当期純利益（損失）",        "当期純利益（損失）",        "total",    True),
]

SGA_TMPL = [
    ("給料・賞与",             "従業員給料・賞与",          "item",     False),
    ("法定福利費",             "法定福利費",              "item",     False),
    ("福利厚生費",             "福利厚生費",              "item",     False),
    ("退職給付費用",            "退職給付費用",            "item",     False),
    ("減価償却費",             "減価償却費",              "item",     False),
    ("租税公課",              "租税公課",               "item",     False),
    ("賃借料",               "賃借料",                 "item",     False),
    ("賞与引当金繰入",          "賞与引当金繰入",           "item",     False),
    ("旅費交通費",             "旅費交通費",              "item",     False),
    ("広告宣伝費",             "広告宣伝費",              "item",     False),
    ("販売促進費",             "販売促進費",              "item",     False),
    ("発送配達費",             "発送配達費",              "item",     False),
    ("車両費",               "車両費",                 "item",     False),
    ("業務委託費",             "業務委託費",              "item",     False),
    ("支払手数料",             "支払手数料",              "item",     False),
    ("リース料",              "リース料",               "item",     False),
    ("地代家賃",              "地代家賃",               "item",     False),
    ("事務用消耗品費",          "事務用消耗品費",           "item",     False),
    ("通信費",               "通信費",                 "item",     False),
    ("水道光熱費",             "水道光熱費",              "item",     False),
    ("寄付金",               "寄付金",                 "item",     False),
    ("接待交際費",             "接待交際費",              "item",     False),
    ("保険料",               "保険料",                 "item",     False),
    ("備品消耗品費",            "備品消耗品費",            "item",     False),
    ("管理諸費",              "管理諸費",               "item",     False),
    ("会議費",               "会議費",                 "item",     False),
    ("研究開発費",             "研究開発費",              "item",     False),
    ("その他販管費",            "その他販管費",            "item",     False),
    ("販売費及び一般管理費合計",   "販売費及び一般管理費合計",   "total",    True),
]

# ── 業種→銘柄マッピング ─────────────────────────
INDUSTRY_TICKERS = {
    "製造":    ["7203.T","7267.T","7201.T","7269.T","6981.T","6762.T","6954.T"],
    "自動車":  ["7203.T","7267.T","7201.T","7269.T","7270.T","7261.T"],
    "電子部品":["6981.T","6762.T","6594.T","6723.T","6963.T","6857.T"],
    "機械":    ["6954.T","6861.T","6113.T","6103.T","7011.T","6301.T"],
    "食品":    ["2802.T","2801.T","2897.T","2503.T","2269.T","2871.T"],
    "化学":    ["4063.T","3407.T","4005.T","4183.T","4911.T","4901.T"],
    "鉄鋼":    ["5401.T","5411.T","5801.T","5706.T"],
    "建設":    ["1812.T","1803.T","1801.T","1802.T","1821.T","1925.T"],
    "不動産":  ["8801.T","8802.T","8830.T","8804.T","3289.T"],
    "IT":      ["4307.T","3659.T","4689.T","6702.T","6701.T"],
    "情報":    ["4307.T","3659.T","6702.T","6701.T"],
    "通信":    ["9432.T","9433.T","9434.T","9984.T"],
    "小売":    ["8267.T","3382.T","9983.T","2651.T","3099.T"],
    "飲食":    ["7550.T","8179.T","9861.T","3197.T","3543.T"],
    "サービス":["4661.T","2432.T","6098.T","2413.T","4755.T"],
    "医療":    ["4502.T","4503.T","4506.T","4507.T","7741.T"],
    "物流":    ["9064.T","9147.T","9076.T","9057.T"],
    "運輸":    ["9020.T","9022.T","9001.T","9202.T","9201.T"],
}

INDUSTRY_ALIAS = {
    "seizo": "製造", "manufacturing": "製造", "mfg": "製造",
    "it": "IT", "software": "IT", "tech": "IT",
    "service": "サービス", "services": "サービス",
    "kensetsu": "建設", "construction": "建設",
    "shokuhin": "食品", "food": "食品",
    "inshoku": "飲食", "restaurant": "飲食",
    "iryo": "医療", "medical": "医療",
    "unyu": "運輸", "transport": "運輸",
    "kouri": "小売", "retail": "小売",
    "fudosan": "不動産", "realestate": "不動産",
    "jidosha": "自動車", "auto": "自動車",
    "kagaku": "化学", "chemical": "化学",
    "buturyu": "物流", "logistics": "物流",
}

def get_tickers(keyword):
    kw = INDUSTRY_ALIAS.get(keyword.lower(), keyword)
    for key, tickers in INDUSTRY_TICKERS.items():
        if key in kw or kw in key:
            return tickers
    return INDUSTRY_TICKERS["製造"]

# ── ユーティリティ ──────────────────────────────
def to_int(s):
    if s is None: return None
    s = str(s).strip().replace(',','').replace(' ','').replace('　','')
    s = s.replace('△','-').replace('▲','-')
    try: return int(float(s))
    except: return None

def is_num(s):
    s = str(s).strip().replace(',','').replace(' ','')
    return bool(re.match(r'^-?\d+(\.\d+)?$', s))

# ── 法人事業概況説明書 パーサー ──────────────────
def _normalize_digits(text):
    """空白区切り1桁数字列を連結: "9 1 1 6 3 0 0 0" → "91163000" """
    text = re.sub(r'(-)\s+(\d(?:\s+\d)*)',
                  lambda m: m.group(1) + m.group(2).replace(' ', ''), text)
    text = re.sub(r'(?<!\d)(\d)(\s\d){1,10}(?!\d)',
                  lambda m: m.group(0).replace(' ', ''), text)
    return text

def _fv(text, *patterns):
    for pat in patterns:
        m = re.search(pat, text, re.DOTALL)
        if m:
            s = m.group(1).replace(' ', '').replace(',', '')
            try: return int(float(s))
            except: pass
    return None

def is_jigyoukyo_page(page):
    text = page.extract_text() or ''
    return '法人事業概況説明書' in text or '各科目の単位：円' in text

def parse_jigyoukyo(page):
    """法人事業概況説明書からPL・BS値を抽出"""
    raw  = page.extract_text() or ''
    norm = _normalize_digits(raw)

    sales  = _fv(norm, r'単位[：:]\s*円\s+(-?\d{4,})')
    cogs   = _fv(norm, r'収\s*入\s*[）)]\s*原\s*価\s+(-?\d+)')
    gross  = _fv(norm, r'売上（収入）総利益\s+(-?\d+)')
    op     = _fv(norm, r'営\s*業\s*損\s*益\s+(-?\d+)')
    pretax = _fv(norm, r'税\s*引\s*前\s*当\s*期\s*損\s*益\s+(-?\d+)')
    assets = _fv(norm, r'資\s*産\s*の\s*部\s*合\s*計\s+(-?\d+)')
    cash   = _fv(norm, r'現\s*金\s*預\s*金\s+(-?\d+)')
    recv_m = re.search(r'注[0-9０-９]\s*\n\s*(-?\d{4,})', norm, re.DOTALL)
    recv   = int(recv_m.group(1)) if recv_m else None
    equity = _fv(norm, r'純\s*資\s*産\s*の\s*部\s*合\s*計\s+(-?\d+)')
    depr   = _fv(norm, r'償\s*却\s*費\s+(-?\d{4,})')
    salary = _fv(norm, r'従\s*業\s*員\s*給\s*料\s+(-?\d+)')
    rent   = _fv(norm, r'地\s*代\s*家\s*賃\s+(-?\d+)')
    liab   = (assets - equity) if (assets is not None and equity is not None) else None
    # 売上原価内訳
    koba_s = _fv(norm, r'期\s*首\s*棚\s*卸\s*高\s+(-?\d+)')
    shii   = _fv(norm, r'材\s*料\s*費\s*（\s*仕\s*入\s*高\s*）\s+(-?\d+)',
                        r'当\s*期\s*仕\s*入\s*高\s+(-?\d+)')
    koba_e = _fv(norm, r'期\s*末\s*棚\s*卸\s*高\s+(-?\d+)')
    # 販管費追加
    kosai  = _fv(norm, r'交\s*際\s*費\s+(-?\d+)')
    # 買掛金（注番号が挟まるケースに対応）
    _km = (re.search(r'買\s*\n\s*注[0-9０-９]\s*\n\s*掛\s*金\s+(-?\d+)', norm) or
           re.search(r'買\s*掛\s*金\s+(-?\d+)', norm))
    kaikake = int(_km.group(1).replace(' ', '').replace(',', '')) if _km else None
    # 借入金（個人＝短期、その他＝長期として近似）
    kojin_kari = _fv(norm, r'個\s*人\s*借\s*入\s*金\s+(-?\d+)')
    other_kari = _fv(norm, r'そ\s*の\s*他\s*借\s*入\s*金\s+(-?\d+)')

    pl  = {}
    bsa = {}
    bsl = {}
    sga = {}
    if sales  is not None: pl['売上高']                   = sales
    if koba_s is not None: pl['期首棚卸高']               = koba_s
    if shii   is not None: pl['当期仕入高']               = shii
    if koba_e is not None: pl['期末棚卸高']               = koba_e
    if cogs   is not None: pl['売上原価合計']              = cogs
    if gross  is not None: pl['売上総利益（損失）']        = gross
    if op     is not None: pl['営業利益（損失）']          = op
    if pretax is not None: pl['税引前当期純利益（損失）']  = pretax
    if assets is not None: bsa['資産合計']                = assets
    if cash   is not None: bsa['現金及び預金']            = cash
    if recv   is not None: bsa['売掛金']                  = recv
    if liab        is not None: bsl['負債合計']             = liab
    if kaikake     is not None: bsl['買掛金']              = kaikake
    if kojin_kari  is not None: bsl['短期借入金']          = kojin_kari
    if other_kari  is not None: bsl['長期借入金']          = other_kari
    if equity      is not None: bsl['純資産合計']          = equity
    if assets is not None: bsl['負債・純資産合計']        = assets
    if depr   is not None: sga['減価償却費']              = depr
    if salary is not None: sga['従業員給料・賞与']        = salary
    if rent   is not None: sga['地代家賃']                = rent
    if kosai  is not None: sga['接待交際費']              = kosai
    return pl, bsa, bsl, sga

# ── PDF解析（位置ベース）───────────────────────
def get_num_rows(page, x_split=300, min_y=115):
    """ページから数値を行別・左右別に抽出（ヘッダー行除外）"""
    try:
        words = page.extract_words(keep_blank_chars=True, x_tolerance=4, y_tolerance=4)
    except Exception:
        return {}
    rows = {}
    for w in words:
        if not is_num(w['text'].replace(',','')):
            continue
        val = to_int(w['text'])
        if val is None:
            continue
        y = round(float(w['top']) / 5) * 5
        if y < min_y:
            continue
        side = 'L' if float(w['x0']) < x_split else 'R'
        rows.setdefault(y, {}).setdefault(side, []).append((float(w['x0']), val))
    result = {}
    for y, sides in rows.items():
        result[y] = {}
        for side, items in sides.items():
            items.sort(key=lambda t: t[0])
            result[y][side] = [v for _, v in items]
    return result

def classify_page(page):
    """ページ種別を数値配置パターンで判定"""
    rows = get_num_rows(page, x_split=300)
    if not rows: return None
    n_left  = sum(1 for r in rows.values() if 'L' in r)
    n_right = sum(1 for r in rows.values() if 'R' in r)
    total   = len(rows)
    if n_left >= 5 and n_right >= 5 and total <= 22: return 'bs'
    if n_left <= 4 and n_right >= 8 and total <= 18: return 'pl'
    if n_right >= 18 and total <= 35: return 'sga'
    return None

def extract_column_vals(rows, side):
    """指定列の値を y順に返す"""
    vals = []
    for y in sorted(rows.keys()):
        r = rows[y]
        if side in r:
            vals.append(r[side][-1])
    return vals

def parse_bs(page):
    """BS: アンカー+順次マッピング"""
    rows = get_num_rows(page, x_split=300)
    left_vals  = extract_column_vals(rows, 'L')
    right_vals = extract_column_vals(rows, 'R')

    asset_data = {}
    liab_data  = {}

    # ── 資産側（左列） ──
    # アンカー: 最初=流動資産合計, 最後=資産合計
    if left_vals:
        asset_data['流動資産合計'] = left_vals[0]
        asset_data['資産合計']    = left_vals[-1]
        # 固定資産合計を探す（資産合計-流動資産合計に近い値）
        target_fixed = left_vals[-1] - left_vals[0]
        fixed_idx = None
        for i, v in enumerate(left_vals[1:-1], 1):
            if target_fixed != 0 and abs(v - target_fixed) / abs(target_fixed) < 0.02:
                fixed_idx = i; break
        if fixed_idx:
            asset_data['固定資産合計'] = left_vals[fixed_idx]
            current_details = left_vals[1:fixed_idx]
            fixed_details   = left_vals[fixed_idx+1:-1]
        else:
            current_details = left_vals[1:max(1, len(left_vals)//2)]
            fixed_details   = left_vals[max(1, len(left_vals)//2):-1]

        # 流動資産明細を一般的な科目名に割り当て
        current_names = ["現金及び預金","受取手形","売掛金","商品・製品",
                         "仕掛品","原材料・貯蔵品","前払費用","未収入金","その他流動資産"]
        for i, v in enumerate(current_details):
            k = current_names[i] if i < len(current_names) else f"流動資産その他{i+1}"
            asset_data[k] = v

        # 固定資産明細を割り当て
        fixed_names = ["有形固定資産計","建物・構築物","　車両運搬具","　減価償却累計額",
                       "無形固定資産合計","投資その他資産合計","敷金・保証金","繰延税金資産","その他投資資産"]
        for i, v in enumerate(fixed_details):
            k = fixed_names[i] if i < len(fixed_names) else f"固定資産その他{i+1}"
            asset_data[k] = v

    # ── 負債・純資産側（右列） ──
    # アンカー: 最初=流動負債合計, 最後=負債・純資産合計, 最後-1=純資産合計
    if right_vals:
        liab_data['流動負債合計']    = right_vals[0]
        liab_data['負債・純資産合計'] = right_vals[-1]
        if len(right_vals) >= 2:
            liab_data['純資産合計'] = right_vals[-2]

        # 負債合計を探す（流動負債合計と同値の2回目の出現 = 固定負債なしのケース）
        end_idx = None
        for i, v in enumerate(right_vals[1:-2], 1):
            if v == right_vals[0]:
                liab_data['負債合計'] = v
                end_idx = i
                break
        if end_idx is None and len(right_vals) >= 4:
            liab_data['負債合計'] = right_vals[-3]
            end_idx = len(right_vals) - 3
        if end_idx is None:
            end_idx = len(right_vals) // 2

        # 流動負債明細
        current_details = right_vals[1:end_idx]
        current_liab_names = ["買掛金","支払手形","短期借入金","1年内返済長期借入金",
                              "未払金","未払費用","未払法人税等","前受金","預り金","賞与引当金","その他流動負債"]
        for i, v in enumerate(current_details):
            k = current_liab_names[i] if i < len(current_liab_names) else f"流動負債その他{i+1}"
            liab_data[k] = v

        # 純資産明細（負債合計以降、純資産合計・負債純資産合計の前）
        equity_details = right_vals[end_idx+1:-2]
        equity_names = ["株主資本合計","資本金","資本剰余金","利益剰余金合計",
                        "繰越利益剰余金","自己株式","その他包括利益累計額合計"]
        for i, v in enumerate(equity_details):
            k = equity_names[i] if i < len(equity_names) else f"純資産その他{i+1}"
            liab_data[k] = v

    return asset_data, liab_data

def parse_pl(page):
    """PL: アンカー+順次マッピング"""
    rows = get_num_rows(page, x_split=450)
    sorted_ys = sorted(rows.keys())

    # PL_ROW_MAP（第5期KAISEIのフォーマット）
    PL_ROW_MAP = [
        (None,                 "売上高"),
        ("当期製品製造原価",    None),
        (None,                 "売上原価合計"),
        (None,                 "売上総利益（損失）"),
        (None,                 "販売費及び一般管理費合計"),
        (None,                 "営業利益（損失）"),
        ("受取利息",           None),
        ("雑収入",             None),
        ("その他営業外収益",    "営業外収益合計"),
        (None,                 "営業外費用合計"),
        (None,                 "経常利益（損失）"),
        (None,                 "税引前当期純利益（損失）"),
        (None,                 "法人税等"),
        (None,                 "当期純利益（損失）"),
    ]
    pl_data = {}
    for i, y in enumerate(sorted_ys):
        if i >= len(PL_ROW_MAP): break
        ln, rn = PL_ROW_MAP[i]
        r = rows[y]
        if ln and 'L' in r: pl_data[ln] = r['L'][0]
        if rn and 'R' in r: pl_data[rn] = r['R'][-1]

    # フォールバック: アンカー確保
    right_vals = extract_column_vals(rows, 'R')
    if right_vals:
        if '売上高' not in pl_data:             pl_data['売上高'] = right_vals[0]
        if '当期純利益（損失）' not in pl_data: pl_data['当期純利益（損失）'] = right_vals[-1]
        if '経常利益（損失）' not in pl_data and len(right_vals) >= 4:
            pl_data['経常利益（損失）'] = right_vals[-4]
    return pl_data

def parse_sga(page):
    """販管費: 順次マッピング（最後=合計アンカー）"""
    rows = get_num_rows(page, x_split=100)
    sorted_ys = sorted(rows.keys())
    sga_keys = [key for _, key, _, _ in SGA_TMPL]  # キー名で保存（ラベル名ではなく）

    sga_data = {}
    # 最後の行は合計行なのでループから除外し、詳細行のみをテンプレートに割り当て
    detail_ys = sorted_ys[:-1] if len(sorted_ys) > 1 else sorted_ys
    for i, y in enumerate(detail_ys):
        r = rows[y]
        vals = r.get('R', r.get('L', []))
        if not vals: continue
        val = vals[-1]
        if i < len(sga_keys):
            sga_data[sga_keys[i]] = val

    # 合計アンカー: 最後の行の値を販管費合計に確定
    if sorted_ys:
        last_row = rows[sorted_ys[-1]]
        last_vals = last_row.get('R') or last_row.get('L') or []
        if last_vals:
            sga_data['販売費及び一般管理費合計'] = last_vals[-1]

    return sga_data

def extract_pdf(pdf_path):
    result = {'period': Path(pdf_path).stem,
              'bs_asset': {}, 'bs_liab': {}, 'pl': {}, 'sga': {}}
    with pdfplumber.open(str(pdf_path)) as pdf:
        total_words = sum(len(p.extract_words() or []) for p in pdf.pages)
        if total_words == 0:
            print(f'  [警告] スキャン画像PDFのため読み取り不可: {Path(pdf_path).name}')
            print(f'         テキスト形式のPDFに変換してください。')
            return result

        # 法人事業概況説明書が含まれる確定申告書形式か判定
        jigyoukyo_pages = [p for p in pdf.pages if is_jigyoukyo_page(p)]

        if jigyoukyo_pages:
            # 確定申告書形式: 法人事業概況説明書ページのみ解析（他ページは無視）
            for page in jigyoukyo_pages:
                pl, bsa, bsl, sga = parse_jigyoukyo(page)
                result['pl'].update(pl)
                result['bs_asset'].update(bsa)
                result['bs_liab'].update(bsl)
                result['sga'].update(sga)
        else:
            # 通常形式: 位置ベース解析（BS/PL形式のPDFのみ対象）
            all_text = ' '.join(p.extract_text() or '' for p in pdf.pages)
            has_bs_pl = '貸借対照表' in all_text or '損益計算書' in all_text
            if not has_bs_pl:
                print(f'  [警告] 消費税申告書など非BS/PL形式のためスキップ: {Path(pdf_path).name}')
            else:
                for page in pdf.pages:
                    ptype = classify_page(page)
                    if ptype == 'bs':
                        a, l = parse_bs(page)
                        result['bs_asset'].update(a)
                        result['bs_liab'].update(l)
                    elif ptype == 'pl':
                        result['pl'].update(parse_pl(page))
                    elif ptype == 'sga':
                        result['sga'].update(parse_sga(page))

    if not result['bs_asset'] and not result['pl']:
        print(f'  [警告] BS・PLが認識できません: {Path(pdf_path).name}')
        print(f'         消費税申告書のみ・試算表形式などでは財務データを取得できません。')
    return result

# ── 同業他社マルチプル ──────────────────────────
def fetch_multiples(industry):
    tickers = get_tickers(industry)
    print(f"  同業他社データ取得中 ({industry}: {len(tickers)}社)...")
    rows = []
    for ticker in tickers:
        try:
            info = yf.Ticker(ticker).info
            name = info.get('longName') or info.get('shortName', ticker)
            per = info.get('trailingPE'); pbr = info.get('priceToBook')
            ev_ebitda = info.get('enterpriseToEbitda'); ev_sales = info.get('enterpriseToRevenue')
            if any(v is not None for v in [per, pbr, ev_ebitda, ev_sales]):
                rows.append({'会社名': name, 'PER': per, 'PBR': pbr,
                             'EV/EBITDA': ev_ebitda, 'EV/売上': ev_sales})
        except Exception:
            pass
    if not rows: return None, {}
    df = pd.DataFrame(rows)
    medians = {}
    for col in ['PER', 'PBR', 'EV/EBITDA', 'EV/売上']:
        vals = df[col].dropna(); vals = vals[(vals > 0) & (vals < 500)]
        if len(vals): medians[col] = round(float(vals.median()), 2)
    return df, medians

# ── Excelスタイルヘルパー ───────────────────────
def bdr():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def sty(cell, bold=False, fg='000000', bg=None, size=10,
        ha='left', fmt=None, wrap=False, italic=False):
    cell.font = Font(bold=bold, italic=italic, size=size, color=fg)
    if bg: cell.fill = PatternFill(fill_type='solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cell.border = bdr()
    if fmt: cell.number_format = fmt

def write_title(ws, row, text, n=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    c = ws.cell(row=row, column=1, value=text)
    sty(c, bold=True, size=12, fg='FFFFFF', bg=C_HEADER, ha='center')
    ws.row_dimensions[row].height = 24

def write_notice(ws, row, n=6):
    """凡例行"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    c = ws.cell(row=row, column=1,
        value='【凡例】緑=合計（確定）/ 水色=小計（確定）/ 黄色=自動入力（要確認・科目名を修正してください）/ 灰色=未入力（手動入力）')
    c.font = Font(size=8, color='666666', italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 14

def write_hdr(ws, row, periods):
    p = periods
    hdrs = ["科目",
            p[0] if len(p)>0 else "第X期",
            p[1] if len(p)>1 else "第X+1期",
            p[2] if len(p)>2 else "第X+2期",
            "増減額\n(2→3期)", "増減率\n(2→3期)"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        sty(c, bold=True, fg='FFFFFF', bg=C_MID, ha='center', wrap=True)
    ws.row_dimensions[row].height = 30

def set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'B4'

def gv(all_data, key, sec):
    return [d[sec].get(key) for d in all_data]

def write_data_row(ws, row, label, vals, rtype, confirmed):
    """
    rtype: header/item/subtotal/total
    confirmed: True=確定(合計等), False=要確認(自動入力明細)
    """
    lc = ws.cell(row=row, column=1, value=label)
    if rtype == 'header':
        sty(lc, bold=True, fg='FFFFFF', bg='4472C4')
        for col in range(2, 7): sty(ws.cell(row=row, column=col), bg='4472C4')
        ws.row_dimensions[row].height = 18
        return

    any_val = any(v is not None for v in vals[:3])

    if rtype == 'total':
        bg = C_TOTAL; bold = True
    elif rtype == 'subtotal':
        bg = C_SUBTOTAL; bold = True
    elif any_val and not confirmed:
        bg = C_AUTO; bold = False   # 黄色: 自動入力・要確認
    elif any_val and confirmed:
        bg = None; bold = False
    else:
        bg = C_EMPTY; bold = False  # 灰色: 未入力テンプレート

    # 科目ラベル（要確認は斜体で表示）
    sty(lc, bold=bold, bg=bg, italic=(not confirmed and any_val and rtype=='item'))

    v2 = vals[1] if len(vals)>1 else None
    v3 = vals[2] if len(vals)>2 else None

    for ci, v in enumerate(vals[:3], 2):
        c = ws.cell(row=row, column=ci)
        if v is not None:
            c.value = v
            sty(c, bold=bold, bg=bg, ha='right', fmt='#,##0;[Red]-#,##0')
        else:
            c.value = ''
            sty(c, bold=bold, bg=bg, ha='center')

    # 増減額・増減率
    cd = ws.cell(row=row, column=5)
    cr = ws.cell(row=row, column=6)
    if v2 is not None and v3 is not None:
        diff = v3 - v2
        cd.value = diff
        sty(cd, bold=bold, bg=bg, ha='right', fmt='#,##0;[Red]-#,##0')
        if v2 != 0:
            cr.value = diff / abs(v2)
            sty(cr, bold=bold, bg=bg, ha='right', fmt='0.0%;[Red]-0.0%')
        else:
            cr.value = '-'; sty(cr, bold=bold, bg=bg, ha='center')
    else:
        cd.value = ''; sty(cd, bold=bold, bg=bg, ha='center')
        cr.value = ''; sty(cr, bold=bold, bg=bg, ha='center')

# ── BS出力 ─────────────────────────────────────
def write_bs(ws, all_data, periods):
    write_title(ws, 1, '貸借対照表　（単位：円）')
    write_notice(ws, 2)
    write_hdr(ws, 3, periods)
    set_cols(ws, [30,14,14,14,14,10])
    row = 4
    for label, key, rtype, confirmed in BS_ASSET_TMPL:
        if key is None:
            write_data_row(ws, row, label, [], 'header', True)
        else:
            vals = gv(all_data, key, 'bs_asset')
            write_data_row(ws, row, label, vals, rtype, confirmed)
        row += 1
    row += 1
    for label, key, rtype, confirmed in BS_LIAB_TMPL:
        if key is None:
            write_data_row(ws, row, label, [], 'header', True)
        else:
            vals = gv(all_data, key, 'bs_liab')
            write_data_row(ws, row, label, vals, rtype, confirmed)
        row += 1

# ── PL出力 ─────────────────────────────────────
def write_pl(ws, all_data, periods):
    write_title(ws, 1, '損益計算書　（単位：円）')
    write_notice(ws, 2)
    write_hdr(ws, 3, periods)
    set_cols(ws, [34,14,14,14,14,10])
    row = 4
    for label, key, rtype, confirmed in PL_TMPL:
        if key is None:
            write_data_row(ws, row, label, [], 'header', True)
        else:
            vals = gv(all_data, key, 'pl')
            write_data_row(ws, row, label, vals, rtype, confirmed)
        row += 1

# ── 販管費出力 ──────────────────────────────────
def write_sga(ws, all_data, periods):
    write_title(ws, 1, '販売費及び一般管理費　（単位：円）')
    write_notice(ws, 2)
    write_hdr(ws, 3, periods)
    set_cols(ws, [28,14,14,14,14,10])
    row = 4
    for label, key, rtype, confirmed in SGA_TMPL:
        vals = gv(all_data, key, 'sga')
        indent = 0 if rtype == 'total' else 1
        write_data_row(ws, row, '　'*indent + label, vals, rtype, confirmed)
        row += 1

# ── 企業価値算定シート ──────────────────────────
def write_valuation(ws, all_data, comp_df, medians, industry):
    # データがある最新期を使用（スキャンPDF等でデータが空の場合は遡る）
    latest = all_data[-1]
    for d in reversed(all_data):
        if d['bs_liab'].get('純資産合計') is not None or d['pl'].get('売上高') is not None:
            latest = d
            break
    pname  = latest.get('period', '最新期')

    net_assets  = latest['bs_liab'].get('純資産合計')
    op_profit   = latest['pl'].get('営業利益（損失）')
    depreciation= latest['sga'].get('減価償却費', 0) or 0
    net_income  = latest['pl'].get('当期純利益（損失）')
    sales       = latest['pl'].get('売上高')
    ebitda      = (op_profit + depreciation) if op_profit is not None else None

    for col, w in zip('ABCDEF', [32,20,20,22,26,14]):
        ws.column_dimensions[col].width = w
    r = 1

    write_title(ws, r, f'企業価値算定　（対象期: {pname}）')
    ws.row_dimensions[r].height = 26; r += 2

    def sec(label):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=label)
        sty(c, bold=True, fg='FFFFFF', bg=C_MID)
        ws.row_dimensions[r].height = 20; r += 1

    def row2(label, val):
        nonlocal r
        sty(ws.cell(row=r, column=1, value=label))
        c = ws.cell(row=r, column=2, value=val)
        sty(c, ha='right', fmt='#,##0;[Red]-#,##0')
        for col in range(3,7): ws.cell(row=r, column=col).border = bdr()
        r += 1

    sec('■ 基礎財務数値　（単位：円）')
    row2('純資産', net_assets); row2('営業利益（損失）', op_profit)
    row2('減価償却費（販管費内）', depreciation)
    row2('EBITDA（営業利益＋減価償却）', ebitda)
    row2('当期純利益（損失）', net_income); row2('売上高', sales)
    r += 1

    sec('■ 純資産法（修正純資産法）　×2〜4倍レンジ')
    for col, h in enumerate(['算定根拠','×2倍','×3倍','×4倍','算定レンジ',''], 1):
        c = ws.cell(row=r, column=col, value=h)
        sty(c, bold=True, fg='FFFFFF', bg='4472C4', ha='center')
    r += 1
    c = ws.cell(row=r, column=1, value='純資産 × 倍率')
    sty(c, bold=True, bg=C_TOTAL)
    for i, mult in enumerate([2,3,4], 2):
        cv = ws.cell(row=r, column=i)
        cv.value = (net_assets*mult) if net_assets else '-'
        sty(cv, bold=True, bg=C_TOTAL, ha='right', fmt='#,##0;[Red]-#,##0')
    rng = f'{net_assets*2:,.0f} 〜 {net_assets*4:,.0f} 円' if (net_assets and net_assets>0) else \
          ('純資産マイナスのため参考値' if net_assets else '-')
    c5 = ws.cell(row=r, column=5, value=rng)
    sty(c5, bold=True, bg=C_TOTAL, ha='center')
    ws.cell(row=r, column=6).border = bdr(); r += 2

    sec(f'■ マルチプル法　（比較業種: {industry}）')
    if comp_df is not None and not comp_df.empty:
        for col, h in enumerate(['会社名','PER','PBR','EV/EBITDA','EV/売上高',''], 1):
            sty(ws.cell(row=r, column=col, value=h), bold=True, fg='FFFFFF', bg='4472C4', ha='center')
        r += 1
        for _, rd in comp_df.iterrows():
            for col, (key, fmt, ha) in enumerate(
                [('会社名','@','left'),('PER','0.0','right'),('PBR','0.00','right'),
                 ('EV/EBITDA','0.0','right'),('EV/売上','0.00','right')], 1):
                v = rd.get(key); ok = v is not None and pd.notna(v)
                c = ws.cell(row=r, column=col); c.value = v if ok else '-'
                sty(c, ha=ha, fmt=fmt if ok else None)
            ws.cell(row=r, column=6).border = bdr(); r += 1
        mv = ['中央値（採用値）', medians.get('PER'), medians.get('PBR'),
              medians.get('EV/EBITDA'), medians.get('EV/売上')]
        for col, v in enumerate(mv, 1):
            c = ws.cell(row=r, column=col); c.value = v if v is not None else '-'
            sty(c, bold=True, bg=C_SUBTOTAL, ha='right' if col>1 else 'left',
                fmt='0.00' if col>1 and v is not None else None)
        ws.cell(row=r, column=6).border = bdr(); r += 2
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r,column=1,value='※ データ取得失敗。下表の中央値欄にマルチプルを手動入力してください。').font=Font(size=9,color='FF6600')
        r += 2

    sec('◆ マルチプル法　算定結果　（単位：円）')
    for col, h in enumerate(['指標','中央値マルチプル','使用財務数値（円）','算定企業価値（円）','算定根拠',''], 1):
        sty(ws.cell(row=r, column=col, value=h), bold=True, fg='FFFFFF', bg='375623', ha='center')
    r += 1
    calc = [('EV/EBITDA法', medians.get('EV/EBITDA'), ebitda, 'EBITDA × EV/EBITDA中央値'),
            ('EV/売上高法',  medians.get('EV/売上'),   sales,  '売上高 × EV/売上高中央値'),
            ('PBR法',       medians.get('PBR'),       net_assets, '純資産 × PBR中央値'),
            ('PER法',       medians.get('PER'),       net_income, '当期純利益 × PER中央値')]
    for label, mult, base, basis in calc:
        ev = int(mult*base) if (mult and base) else None
        for col, (val, ha, fmt) in enumerate([
            (label,'left',None),(mult,'right','0.00x'),(base,'right','#,##0;[Red]-#,##0'),
            (ev,'right','#,##0;[Red]-#,##0'),(basis,'left',None)], 1):
            c = ws.cell(row=r, column=col); c.value = val if val is not None else '-'
            sty(c, bold=(col==4), bg=C_TOTAL if col==4 else None, ha=ha, fmt=fmt if val is not None else None)
        ws.cell(row=r, column=6).border = bdr(); r += 1

    r += 1
    for note in ['【注記】',
                 '・純資産法の2〜4倍は中小企業算定の目安。業種・収益力に応じて調整してください。',
                 '・マルチプルはyfinanceの概算値。公式情報で確認してください。',
                 '・赤字企業のPER・EV/EBITDA法はマイナスとなるため参考値として扱ってください。',
                 '・株主価値 = EV - 純有利子負債（借入金 - 現金）']:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=note).font = Font(size=9, color='666666')
        r += 1

# ── メイン ─────────────────────────────────────
def main():
    sys.stdout.reconfigure(encoding='utf-8')
    args      = sys.argv[1:]
    pdf_paths = [a for a in args if a.lower().endswith('.pdf')]
    others    = [a for a in args if not a.lower().endswith('.pdf')]
    industry  = others[0] if others else '製造'

    # PDF未指定の場合: スクリプトと同じフォルダのPDFを自動検出
    script_dir = Path(sys.argv[0]).resolve().parent
    if not pdf_paths:
        # ファイル名中の日付数字でソート（例: 202305, 20240529→202405）して時系列順に
        def _date_key(path):
            m = re.search(r'(\d{6,8})', path.name)
            return m.group(1)[:6] if m else path.name
        found = sorted(script_dir.glob('*.pdf'), key=_date_key)
        if not found:
            print('[エラー] このフォルダにPDFファイルが見つかりません。')
            print(f'  フォルダ: {script_dir}')
            print('  PDFをこのフォルダに入れてから再実行してください。')
            input('Enterキーで終了...')
            sys.exit(1)
        print(f'PDFを検出しました ({len(found)}件):')
        for i, p in enumerate(found, 1):
            print(f'  {i}. {p.name}')
        if len(found) > 3:
            # 財務データがあるPDFを優先して最大3件選択
            print('  ※ 財務データを含むPDFを優先して3件を選択します...')
            def has_financial_data(path):
                try:
                    import pdfplumber as _plumber
                    with _plumber.open(str(path)) as pdf:
                        for page in pdf.pages:
                            if is_jigyoukyo_page(page):
                                return True
                        all_text = ' '.join(p.extract_text() or '' for p in pdf.pages)
                        return '貸借対照表' in all_text or '損益計算書' in all_text
                except Exception:
                    return False
            usable = [p for p in found if has_financial_data(p)]
            if len(usable) >= 3:
                selected = usable[:3]
            else:
                selected = found[:3]
            print(f'  → 使用ファイル:')
            for p in selected:
                print(f'      {p.name}')
            pdf_paths = [str(p) for p in selected]
        else:
            pdf_paths = [str(p) for p in found]
        print()

    print('=== 決算書PDF -> 3期比較表 + 企業価値算定 ===')
    print(f'PDF数: {len(pdf_paths)}  業種: {industry}')

    all_data, periods = [], []
    for path in pdf_paths[:3]:
        print(f'読み込み: {Path(path).name}')
        data = extract_pdf(path)
        all_data.append(data)
        periods.append(data['period'])
        print(f'  BS資産={len(data["bs_asset"])}  BS負債={len(data["bs_liab"])}'
              f'  PL={len(data["pl"])}  販管費={len(data["sga"])}')

    empty = {'period':'','bs_asset':{},'bs_liab':{},'pl':{},'sga':{}}
    while len(all_data) < 3:
        all_data.insert(0, dict(empty)); periods.insert(0, '')

    comp_df, medians = fetch_multiples(industry)
    if medians: print(f'  マルチプル中央値: {medians}')

    print('Excel生成中...')
    wb = Workbook()
    ws_bs  = wb.active;            ws_bs.title = '貸借対照表'
    ws_pl  = wb.create_sheet('損益計算書')
    ws_sga = wb.create_sheet('販売費及び一般管理費')
    ws_val = wb.create_sheet('企業価値算定')

    write_bs(ws_bs, all_data, periods)
    write_pl(ws_pl, all_data, periods)
    write_sga(ws_sga, all_data, periods)
    write_valuation(ws_val, all_data, comp_df, medians, industry)

    output = script_dir / 'kessan_output.xlsx'
    wb.save(str(output))
    print(f'完了: {output}')

if __name__ == '__main__':
    main()
