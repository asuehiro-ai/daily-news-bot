#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
決算分析・企業価値算定ツール生成スクリプト
既存の「3期比較表.xlsx」に以下を追加:
  - 業界平均比較・アラートシート
  - 企業価値算定（修正前）
  - 企業価値算定（修正後）
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import shutil, os, sys

# ─────────────────────────────────────────────
# セル位置定数（3期比較表.xlsx の構造に基づく）
# 列: A=科目, B=第1期, C=第2期, D=第3期
# ─────────────────────────────────────────────
BS  = '貸借対照表'
PL  = '損益計算書'
SGA = '販売費及び一般管理費'

# 貸借対照表 行番号
BS_R = dict(
    流動資産合計=14, 固定資産合計=28, 資産合計=29,
    流動負債合計=43, 固定負債合計=48, 負債合計=49,
    純資産合計=60,  負債純資産合計=61,
    有形固定資産計=16, 無形固定資産合計=23,
)
# 損益計算書 行番号
PL_R = dict(
    売上高=4, 売上原価合計=10, 売上総利益=11,
    販管費合計=12, 営業利益=13,
    営業外収益合計=18, 営業外費用合計=23, 経常利益=24,
    税引前当期純利益=34, 法人税等=35, 当期純利益=36,
)
# 販管費 行番号
SGA_R = dict(給料賞与=4, 減価償却費=8, 販管費合計=32)

# 期の列 (B=第1期/第5期, C=第2期/第6期, D=第3期/第7期)
PCOL = {1: 'B', 2: 'C', 3: 'D'}

# ─────────────────────────────────────────────
# スタイル
# ─────────────────────────────────────────────
def mk_font(bold=False, color='262626', size=9, name='游ゴシック'):
    return Font(name=name, size=size, bold=bold, color=color)

def mk_fill(color):
    return PatternFill('solid', fgColor=color)

def mk_border(t='thin', c='BFBFBF'):
    s = Side(style=t, color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def mk_align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# 色定数
C_NAVY  = '1F3864'
C_BLUE  = '2E75B6'
C_LBLUE = 'BDD7EE'
C_LLBLUE= 'DEEAF1'
C_GRAY  = 'F2F2F2'
C_LGRAY = 'FAFAFA'
C_GOLD  = 'FFF2CC'
C_RED   = 'FFC7CE'
C_YEL   = 'FFEB9C'
C_GRN   = 'C6EFCE'
C_INPUT = 'FFFDE7'
C_WHITE = 'FFFFFF'
C_DRED  = '9C0006'
C_DYEL  = '7D4706'
C_DGRN  = '375623'

def fmt_cell(cell, val=None, bold=False, color='262626', fill=None,
             align_h='left', align_v='center', wrap=False,
             border=True, num_fmt=None, size=9):
    if val is not None:
        cell.value = val
    cell.font = mk_font(bold=bold, color=color, size=size)
    if fill:
        cell.fill = mk_fill(fill)
    cell.alignment = mk_align(h=align_h, v=align_v, wrap=wrap)
    if border:
        cell.border = mk_border()
    if num_fmt:
        cell.number_format = num_fmt

def title_row(ws, row, text, col_from=1, col_to=10, height=28):
    ws.row_dimensions[row].height = height
    cell = ws.cell(row=row, column=col_from)
    cell.value = text
    cell.font = mk_font(bold=True, color=C_WHITE, size=13)
    cell.fill = mk_fill(C_NAVY)
    cell.alignment = mk_align(h='left', v='center')
    cell.border = mk_border()
    if col_to > col_from:
        ws.merge_cells(start_row=row, start_column=col_from,
                       end_row=row, end_column=col_to)

def section_row(ws, row, text, col_from=1, col_to=10, height=20):
    ws.row_dimensions[row].height = height
    cell = ws.cell(row=row, column=col_from)
    cell.value = text
    cell.font = mk_font(bold=True, color=C_NAVY, size=9)
    cell.fill = mk_fill(C_LBLUE)
    cell.alignment = mk_align(h='left', v='center')
    cell.border = mk_border()
    if col_to > col_from:
        ws.merge_cells(start_row=row, start_column=col_from,
                       end_row=row, end_column=col_to)

def hdr_cell(cell, text, align_h='center'):
    cell.value = text
    cell.font = mk_font(bold=True, color=C_WHITE, size=9)
    cell.fill = mk_fill(C_BLUE)
    cell.alignment = mk_align(h=align_h, v='center', wrap=True)
    cell.border = mk_border()

def data_cell(cell, val=None, bold=False, fill=None, align_h='right',
              num_fmt='#,##0', formula=False):
    if val is not None:
        cell.value = val
    cell.font = mk_font(bold=bold, color='262626')
    if fill:
        cell.fill = mk_fill(fill)
    cell.alignment = mk_align(h=align_h)
    cell.border = mk_border()
    if num_fmt and not formula:
        cell.number_format = num_fmt

def label_cell(cell, text, bold=False, fill=None, indent=0):
    cell.value = ('　' * indent) + text
    cell.font = mk_font(bold=bold)
    if fill:
        cell.fill = mk_fill(fill)
    cell.alignment = mk_align(h='left')
    cell.border = mk_border()

def formula_cell(cell, formula, num_fmt='#,##0', fill=None, bold=False,
                 align_h='right'):
    cell.value = formula
    cell.font = mk_font(bold=bold)
    if fill:
        cell.fill = mk_fill(fill)
    cell.alignment = mk_align(h=align_h)
    cell.border = mk_border()
    cell.number_format = num_fmt

def input_cell(cell, hint='', num_fmt='#,##0'):
    cell.value = hint if hint else None
    cell.fill = mk_fill(C_INPUT)
    cell.font = mk_font(color='1F497D')
    cell.alignment = mk_align(h='right')
    cell.border = mk_border()
    cell.number_format = num_fmt

# BS/PL への参照式を作る
def bs(row, period=3):
    return f"=IFERROR('{BS}'!{PCOL[period]}{BS_R[row]},0)"

def pl(row, period=3):
    return f"=IFERROR('{PL}'!{PCOL[period]}{PL_R[row]},0)"

def sga(row, period=3):
    return f"=IFERROR('{SGA}'!{PCOL[period]}{SGA_R[row]},0)"

# ─────────────────────────────────────────────
# 業界平均データ
# ─────────────────────────────────────────────
INDUSTRIES = [
    'IT・情報通信',
    'サービス業',
    '製造業',
    '卸売業',
    '小売業',
    '建設業',
    '飲食業',
    '不動産業',
]

# 各指標の業界平均値 (業種 × 指標)
# 指標: 売総率,営業率,経常率,純利率,自己資本率,流動比率,固定比率,人件費率,ROA,ROE,EV_EBITDA,EV_Sales,PBR,PER
IND_DATA = {
    'IT・情報通信': [0.65, 0.08, 0.09, 0.06, 0.45, 1.60, 0.40, 0.45, 0.06, 0.12, 12.0, 2.5,  2.87, 23.42],
    'サービス業':  [0.60, 0.05, 0.06, 0.04, 0.35, 1.40, 0.60, 0.40, 0.04, 0.10,  8.0, 1.5,  1.50, 15.0 ],
    '製造業':      [0.25, 0.03, 0.04, 0.02, 0.40, 1.30, 0.90, 0.18, 0.03, 0.07,  7.0, 0.8,  1.20, 12.0 ],
    '卸売業':      [0.20, 0.02, 0.03, 0.02, 0.35, 1.20, 0.60, 0.10, 0.02, 0.06,  6.0, 0.3,  1.00, 10.0 ],
    '小売業':      [0.30, 0.02, 0.03, 0.02, 0.30, 1.10, 0.80, 0.15, 0.02, 0.07,  6.0, 0.5,  1.10, 12.0 ],
    '建設業':      [0.22, 0.03, 0.04, 0.02, 0.38, 1.30, 0.60, 0.20, 0.03, 0.08,  5.0, 0.4,  1.00, 10.0 ],
    '飲食業':      [0.65, 0.03, 0.04, 0.02, 0.15, 0.80, 1.20, 0.35, 0.02, 0.10,  5.0, 0.6,  1.50, 15.0 ],
    '不動産業':    [0.55, 0.15, 0.16, 0.10, 0.40, 1.20, 1.50, 0.10, 0.05, 0.12, 10.0, 2.0,  1.80, 18.0 ],
}

IND_COLS = ['売上総利益率','営業利益率','経常利益率','純利益率',
            '自己資本比率','流動比率','固定比率','人件費率',
            'ROA','ROE','EV/EBITDA','EV/売上高','PBR','PER']

# ─────────────────────────────────────────────
# シート作成: 業界データ（非表示）
# ─────────────────────────────────────────────
def create_industry_data_sheet(wb):
    ws = wb.create_sheet('業界データ')
    ws.sheet_state = 'hidden'

    # ヘッダー行: 行1=業種名, 列2~=指標
    ws.cell(1, 1).value = '業種'
    for ci, col in enumerate(IND_COLS, 2):
        ws.cell(1, ci).value = col

    for ri, ind in enumerate(INDUSTRIES, 2):
        ws.cell(ri, 1).value = ind
        for ci, val in enumerate(IND_DATA[ind], 2):
            ws.cell(ri, ci).value = val

# ─────────────────────────────────────────────
# シート作成: 業界平均比較・アラート
# ─────────────────────────────────────────────
def create_industry_comparison_sheet(wb):
    ws_name = '業界平均比較・アラート'
    if ws_name in wb.sheetnames:
        del wb[ws_name]
    ws = wb.create_sheet(ws_name)

    # 列幅
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 22

    # ── タイトル ──
    title_row(ws, 1, '業界平均比較・財務アラート', col_to=9)

    # ── 業種選択 ──
    ws.row_dimensions[2].height = 22
    label_cell(ws.cell(2, 1), '【業種選択】', bold=True, fill=C_GOLD)
    cell_ind = ws.cell(2, 2)
    cell_ind.value = INDUSTRIES[0]
    cell_ind.font = mk_font(bold=True, color=C_NAVY)
    cell_ind.fill = mk_fill(C_INPUT)
    cell_ind.alignment = mk_align(h='center')
    cell_ind.border = mk_border()

    dv = DataValidation(
        type='list',
        formula1=f'"{",".join(INDUSTRIES)}"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.add(cell_ind)
    ws.add_data_validation(dv)

    label_cell(ws.cell(2, 3), '← ここで業種を選択してください', bold=False, fill=C_GOLD)
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=9)

    # 凡例
    ws.row_dimensions[3].height = 18
    label_cell(ws.cell(3, 1), '【判定凡例】', bold=True, fill=C_LGRAY)
    ws.cell(3, 2).value = '● 要注意（業界平均と20%超乖離）'
    ws.cell(3, 2).fill = mk_fill(C_RED)
    ws.cell(3, 2).font = mk_font(bold=True, color=C_DRED)
    ws.cell(3, 2).alignment = mk_align(h='center')
    ws.cell(3, 2).border = mk_border()
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3)
    ws.cell(3, 4).value = '△ 注意（業界平均と10%超乖離）'
    ws.cell(3, 4).fill = mk_fill(C_YEL)
    ws.cell(3, 4).font = mk_font(bold=True, color=C_DYEL)
    ws.cell(3, 4).alignment = mk_align(h='center')
    ws.cell(3, 4).border = mk_border()
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
    ws.cell(3, 6).value = '○ 正常範囲'
    ws.cell(3, 6).fill = mk_fill(C_GRN)
    ws.cell(3, 6).font = mk_font(bold=True, color=C_DGRN)
    ws.cell(3, 6).alignment = mk_align(h='center')
    ws.cell(3, 6).border = mk_border()
    ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)
    ws.cell(3, 8).value = '─ データなし'
    ws.cell(3, 8).fill = mk_fill(C_GRAY)
    ws.cell(3, 8).font = mk_font(color='808080')
    ws.cell(3, 8).alignment = mk_align(h='center')
    ws.cell(3, 8).border = mk_border()
    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=9)

    # ヘッダー行
    ws.row_dimensions[4].height = 36
    headers = ['指標名', '計算式', '第1期実績', '第2期実績', '第3期実績', '業界平均', '差異(3期-平均)', '判定', 'コメント']
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws.cell(4, ci), h, align_h='center')

    # ── 業界平均のVLOOKUP用数式ベース ──
    # B2セルで業種選択、業界データシートからVLOOKUPで取得
    # 業界データ: A2:O9, 列1=業種名, 列2=売上総利益率, ...
    def ind_avg(col_idx):
        # col_idx: 1=売上総利益率, 2=営業利益率, ...
        return f"=IFERROR(VLOOKUP($B$2,業界データ!$A:$O,{col_idx+1},0),\"\")"

    # 期別の数値を取得するセル参照
    def pct_val(sheet_name, row_num, period):
        return f"=IFERROR('{sheet_name}'!{PCOL[period]}{row_num},\"\")"

    # 比率計算式（分子/分母）
    def ratio_formula(num_s, num_r, den_s, den_r, period, is_ratio_not_pct=False):
        n = f"IFERROR('{num_s}'!{PCOL[period]}{num_r},0)"
        d = f"IFERROR('{den_s}'!{PCOL[period]}{den_r},0)"
        return f"=IFERROR({n}/{d},\"\")"

    # アラート判定式（比率指標用）: 差異が絶対値で閾値超かどうか
    def alert_formula(diff_cell, avg_cell, inverse=False):
        # inverse: 大きいほうが悪い(人件費率など)
        sign = -1 if inverse else 1
        return (
            f'=IF({avg_cell}="","─",'
            f'IF(ABS({diff_cell})>ABS({avg_cell})*0.2,"●",'
            f'IF(ABS({diff_cell})>ABS({avg_cell})*0.1,"△","○")))'
        )

    def comment_formula(diff_cell, avg_cell, label, inverse=False):
        hi = f'"{label}が業界平均より大幅に高い"'
        lo = f'"{label}が業界平均より大幅に低い"'
        if inverse:
            hi, lo = lo, hi
        return (
            f'=IF({avg_cell}="","",IF(ABS({diff_cell})>ABS({avg_cell})*0.2,'
            f'IF({diff_cell}>0,{hi},{lo}),"正常範囲"))'
        )

    # ── 収益性分析 ──
    R = 5
    section_row(ws, R, '【収益性分析】', col_to=9)
    ws.row_dimensions[R].height = 18

    ratios_prof = [
        # (指標名, 計算式説明, 分子シート, 分子行, 分母シート, 分母行, 業界列idx, inverse)
        ('売上総利益率', '売上総利益 ÷ 売上高',    PL,'売上総利益', PL,'売上高',    1, False),
        ('営業利益率',   '営業利益 ÷ 売上高',       PL,'営業利益',   PL,'売上高',    2, False),
        ('経常利益率',   '経常利益 ÷ 売上高',       PL,'経常利益',   PL,'売上高',    3, False),
        ('純利益率',     '当期純利益 ÷ 売上高',     PL,'当期純利益', PL,'売上高',    4, False),
        ('人件費率',     '給料賞与 ÷ 売上高',       SGA,'給料賞与',  PL,'売上高',    8, True),
    ]

    for i, (name, formula_desc, ns, nr, ds, dr, ind_ci, inv) in enumerate(ratios_prof):
        row = R + 1 + i
        ws.row_dimensions[row].height = 18

        label_cell(ws.cell(row, 1), name, bold=False, fill=C_WHITE)
        label_cell(ws.cell(row, 2), formula_desc, fill=C_LGRAY)

        for p in [1, 2, 3]:
            c_cell = ws.cell(row, 2 + p)
            n_formula = f"IFERROR('{ns}'!{PCOL[p]}{PL_R[nr] if ns==PL else SGA_R[nr]},0)"
            d_formula = f"IFERROR('{ds}'!{PCOL[p]}{PL_R[dr] if ds==PL else BS_R[dr]},0)"
            c_cell.value = f"=IFERROR(({n_formula})/({d_formula}),\"\")"
            c_cell.number_format = '0.0%'
            c_cell.alignment = mk_align(h='right')
            c_cell.border = mk_border()
            c_cell.font = mk_font()

        avg_cell = ws.cell(row, 6)
        avg_cell.value = ind_avg(ind_ci)
        avg_cell.number_format = '0.0%'
        avg_cell.alignment = mk_align(h='right')
        avg_cell.border = mk_border()
        avg_cell.font = mk_font()
        avg_cell.fill = mk_fill(C_LLBLUE)

        diff_cell = ws.cell(row, 7)
        avg_ref = f'F{row}'
        val_ref = f'E{row}'
        diff_cell.value = f'=IFERROR(IF(OR({val_ref}="",{avg_ref}=""),"",{val_ref}-{avg_ref}),"")'
        diff_cell.number_format = '0.0%'
        diff_cell.alignment = mk_align(h='right')
        diff_cell.border = mk_border()
        diff_cell.font = mk_font()

        judge_cell = ws.cell(row, 8)
        judge_cell.value = alert_formula(f'G{row}', f'F{row}', inverse=inv)
        judge_cell.alignment = mk_align(h='center')
        judge_cell.border = mk_border()
        judge_cell.font = mk_font(bold=True)

        comment = ws.cell(row, 9)
        comment.value = comment_formula(f'G{row}', f'F{row}', name, inverse=inv)
        comment.alignment = mk_align(h='left', wrap=True)
        comment.border = mk_border()
        comment.font = mk_font()

    # ── 安全性分析 ──
    R2 = R + 1 + len(ratios_prof) + 1
    section_row(ws, R2, '【安全性分析】', col_to=9)
    ws.row_dimensions[R2].height = 18

    # 安全性指標は直接計算式を記述
    safety_items = [
        # (名前, 説明, 計算式(C/D/E列), 業界列idx, inverse, 形式)
        ('自己資本比率', '純資産 ÷ 資産合計',
         lambda p: f"=IFERROR('{BS}'!{PCOL[p]}{BS_R['純資産合計']}/'{BS}'!{PCOL[p]}{BS_R['資産合計']},\"\")",
         5, False, '0.0%'),
        ('流動比率', '流動資産 ÷ 流動負債',
         lambda p: f"=IFERROR('{BS}'!{PCOL[p]}{BS_R['流動資産合計']}/IF('{BS}'!{PCOL[p]}{BS_R['流動負債合計']}=0,1,'{BS}'!{PCOL[p]}{BS_R['流動負債合計']}),\"\")",
         6, False, '0.0%'),
        ('固定比率', '固定資産 ÷ 純資産',
         lambda p: f"=IFERROR('{BS}'!{PCOL[p]}{BS_R['固定資産合計']}/IF('{BS}'!{PCOL[p]}{BS_R['純資産合計']}=0,1,'{BS}'!{PCOL[p]}{BS_R['純資産合計']}),\"\")",
         7, True, '0.0%'),
        ('ROA', '経常利益 ÷ 資産合計',
         lambda p: f"=IFERROR('{PL}'!{PCOL[p]}{PL_R['経常利益']}/IF('{BS}'!{PCOL[p]}{BS_R['資産合計']}=0,1,'{BS}'!{PCOL[p]}{BS_R['資産合計']}),\"\")",
         9, False, '0.0%'),
        ('ROE', '当期純利益 ÷ 純資産',
         lambda p: f"=IFERROR('{PL}'!{PCOL[p]}{PL_R['当期純利益']}/IF('{BS}'!{PCOL[p]}{BS_R['純資産合計']}=0,1,'{BS}'!{PCOL[p]}{BS_R['純資産合計']}),\"\")",
         10, False, '0.0%'),
    ]

    for i, (name, desc, f_lambda, ind_ci, inv, nfmt) in enumerate(safety_items):
        row = R2 + 1 + i
        ws.row_dimensions[row].height = 18
        label_cell(ws.cell(row, 1), name, fill=C_WHITE)
        label_cell(ws.cell(row, 2), desc, fill=C_LGRAY)

        for p in [1, 2, 3]:
            c_cell = ws.cell(row, 2 + p)
            c_cell.value = f_lambda(p)
            c_cell.number_format = nfmt
            c_cell.alignment = mk_align(h='right')
            c_cell.border = mk_border()
            c_cell.font = mk_font()

        avg_cell = ws.cell(row, 6)
        avg_cell.value = ind_avg(ind_ci)
        avg_cell.number_format = nfmt
        avg_cell.alignment = mk_align(h='right')
        avg_cell.border = mk_border()
        avg_cell.font = mk_font()
        avg_cell.fill = mk_fill(C_LLBLUE)

        diff_cell = ws.cell(row, 7)
        diff_cell.value = f'=IFERROR(IF(OR(E{row}="",F{row}=""),"",E{row}-F{row}),"")'
        diff_cell.number_format = nfmt
        diff_cell.alignment = mk_align(h='right')
        diff_cell.border = mk_border()
        diff_cell.font = mk_font()

        judge_cell = ws.cell(row, 8)
        judge_cell.value = alert_formula(f'G{row}', f'F{row}', inverse=inv)
        judge_cell.alignment = mk_align(h='center')
        judge_cell.border = mk_border()
        judge_cell.font = mk_font(bold=True)

        comment = ws.cell(row, 9)
        comment.value = comment_formula(f'G{row}', f'F{row}', name, inverse=inv)
        comment.alignment = mk_align(h='left', wrap=True)
        comment.border = mk_border()
        comment.font = mk_font()

    # ── 成長性分析 ──
    R3 = R2 + 1 + len(safety_items) + 1
    section_row(ws, R3, '【成長性分析】', col_to=9)
    ws.row_dimensions[R3].height = 18

    growth_items = [
        ('売上高成長率（1→2期）', '(第2期売上高 - 第1期売上高) ÷ 第1期売上高',
         f"=IFERROR(IF(OR('{PL}'!C{PL_R['売上高']}=\"\",'{PL}'!B{PL_R['売上高']}=0),\"\",('{PL}'!C{PL_R['売上高']}-'{PL}'!B{PL_R['売上高']})/'{PL}'!B{PL_R['売上高']}),\"\")",
         '', 0),
        ('売上高成長率（2→3期）', '(第3期売上高 - 第2期売上高) ÷ 第2期売上高',
         f"=IFERROR(IF(OR('{PL}'!D{PL_R['売上高']}=\"\",'{PL}'!C{PL_R['売上高']}=0),\"\",('{PL}'!D{PL_R['売上高']}-'{PL}'!C{PL_R['売上高']})/'{PL}'!C{PL_R['売上高']}),\"\")",
         11, 0),
    ]

    for i, (name, desc, formula_e, ind_ci, inv) in enumerate(growth_items):
        row = R3 + 1 + i
        ws.row_dimensions[row].height = 18
        label_cell(ws.cell(row, 1), name, fill=C_WHITE)
        label_cell(ws.cell(row, 2), desc, fill=C_LGRAY)

        for p in [1, 2, 3]:
            c_cell = ws.cell(row, 2 + p)
            c_cell.alignment = mk_align(h='right')
            c_cell.border = mk_border()
            c_cell.font = mk_font()
            c_cell.number_format = '0.0%'

        # 期別実績は特別処理（成長率は期の概念が異なる）
        ws.cell(row, 3).value = '─'  # 第1期実績なし
        ws.cell(row, 3).alignment = mk_align(h='center')
        if i == 0:
            ws.cell(row, 4).value = formula_e
            ws.cell(row, 4).number_format = '0.0%'
            ws.cell(row, 4).alignment = mk_align(h='right')
            ws.cell(row, 5).value = '─'
            ws.cell(row, 5).alignment = mk_align(h='center')
        else:
            ws.cell(row, 4).value = '─'
            ws.cell(row, 4).alignment = mk_align(h='center')
            ws.cell(row, 5).value = formula_e
            ws.cell(row, 5).number_format = '0.0%'
            ws.cell(row, 5).alignment = mk_align(h='right')

        avg_cell = ws.cell(row, 6)
        if ind_ci:
            avg_cell.value = ind_avg(ind_ci)
        else:
            avg_cell.value = '─'
        avg_cell.number_format = '0.0%'
        avg_cell.alignment = mk_align(h='right' if ind_ci else 'center')
        avg_cell.border = mk_border()
        avg_cell.font = mk_font()
        avg_cell.fill = mk_fill(C_LLBLUE)

        # 差異・判定
        val_col = 'D' if i == 0 else 'E'
        diff_cell = ws.cell(row, 7)
        diff_cell.value = f'=IFERROR(IF(OR({val_col}{row}="",F{row}="","─"=F{row}),"",{val_col}{row}-F{row}),"")'
        diff_cell.number_format = '0.0%'
        diff_cell.alignment = mk_align(h='right')
        diff_cell.border = mk_border()
        diff_cell.font = mk_font()

        judge_cell = ws.cell(row, 8)
        judge_cell.value = f'=IF(OR(G{row}="","─"=G{row}),"─","○")'
        judge_cell.alignment = mk_align(h='center')
        judge_cell.border = mk_border()
        judge_cell.font = mk_font()

        comment = ws.cell(row, 9)
        comment.value = f'=IF(OR(G{row}="","─"=G{row}),"",IF({val_col}{row}>0,"売上高が成長しています","売上高が減少しています"))'
        comment.alignment = mk_align(h='left')
        comment.border = mk_border()
        comment.font = mk_font()

    # ── 絶対額参照 ──
    R4 = R3 + 1 + len(growth_items) + 1
    section_row(ws, R4, '【主要財務数値（千円）】', col_to=9)
    ws.row_dimensions[R4].height = 18

    abs_items = [
        ('売上高',    f"=IFERROR('{PL}'!{PCOL[1]}{PL_R['売上高']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[2]}{PL_R['売上高']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[3]}{PL_R['売上高']},\"\")"),
        ('売上総利益',f"=IFERROR('{PL}'!{PCOL[1]}{PL_R['売上総利益']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[2]}{PL_R['売上総利益']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[3]}{PL_R['売上総利益']},\"\")"),
        ('営業利益',  f"=IFERROR('{PL}'!{PCOL[1]}{PL_R['営業利益']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[2]}{PL_R['営業利益']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[3]}{PL_R['営業利益']},\"\")"),
        ('経常利益',  f"=IFERROR('{PL}'!{PCOL[1]}{PL_R['経常利益']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[2]}{PL_R['経常利益']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[3]}{PL_R['経常利益']},\"\")"),
        ('当期純利益',f"=IFERROR('{PL}'!{PCOL[1]}{PL_R['当期純利益']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[2]}{PL_R['当期純利益']},\"\")",
                      f"=IFERROR('{PL}'!{PCOL[3]}{PL_R['当期純利益']},\"\")"),
        ('純資産',    f"=IFERROR('{BS}'!{PCOL[1]}{BS_R['純資産合計']},\"\")",
                      f"=IFERROR('{BS}'!{PCOL[2]}{BS_R['純資産合計']},\"\")",
                      f"=IFERROR('{BS}'!{PCOL[3]}{BS_R['純資産合計']},\"\")"),
        ('資産合計',  f"=IFERROR('{BS}'!{PCOL[1]}{BS_R['資産合計']},\"\")",
                      f"=IFERROR('{BS}'!{PCOL[2]}{BS_R['資産合計']},\"\")",
                      f"=IFERROR('{BS}'!{PCOL[3]}{BS_R['資産合計']},\"\")"),
    ]

    for i, (name, f1, f2, f3) in enumerate(abs_items):
        row = R4 + 1 + i
        ws.row_dimensions[row].height = 18
        label_cell(ws.cell(row, 1), name, fill=C_WHITE)
        # B列を計算式説明として使わず、A列見出し、C~Eを数値に変更
        ws.cell(row, 2).value = '（参照）'
        ws.cell(row, 2).font = mk_font(color='888888')
        ws.cell(row, 2).alignment = mk_align(h='center')
        ws.cell(row, 2).border = mk_border()
        ws.cell(row, 2).fill = mk_fill(C_LGRAY)

        for p, fv in enumerate([f1, f2, f3], 3):
            c = ws.cell(row, p)
            c.value = fv
            c.number_format = '#,##0'
            c.alignment = mk_align(h='right')
            c.border = mk_border()
            c.font = mk_font()

        for col in range(6, 10):
            c = ws.cell(row, col)
            c.value = '─'
            c.alignment = mk_align(h='center')
            c.border = mk_border()
            c.font = mk_font(color='808080')
            c.fill = mk_fill(C_GRAY)

    # ── 条件付き書式（H列 判定）──
    last_row = R4 + 1 + len(abs_items)
    red_rule = CellIsRule(operator='equal', formula=['"●"'],
                          fill=mk_fill(C_RED),
                          font=mk_font(bold=True, color=C_DRED))
    yel_rule = CellIsRule(operator='equal', formula=['"△"'],
                          fill=mk_fill(C_YEL),
                          font=mk_font(bold=True, color=C_DYEL))
    grn_rule = CellIsRule(operator='equal', formula=['"○"'],
                          fill=mk_fill(C_GRN),
                          font=mk_font(bold=True, color=C_DGRN))

    range_h = f'H5:H{last_row}'
    ws.conditional_formatting.add(range_h, red_rule)
    ws.conditional_formatting.add(range_h, yel_rule)
    ws.conditional_formatting.add(range_h, grn_rule)

    # G列（差異）に赤・緑条件付き書式
    ws.conditional_formatting.add(f'G5:G{last_row}',
        FormulaRule(formula=[f'AND(NOT(ISBLANK(G5)),G5>0.05)'],
                    fill=mk_fill('FFE0E0'), font=mk_font(color=C_DRED)))
    ws.conditional_formatting.add(f'G5:G{last_row}',
        FormulaRule(formula=[f'AND(NOT(ISBLANK(G5)),G5<-0.05)'],
                    fill=mk_fill('E0FFE0'), font=mk_font(color=C_DGRN)))

    # 注記
    note_row = last_row + 1
    ws.row_dimensions[note_row].height = 60
    note = ws.cell(note_row, 1)
    note.value = (
        '【注記】\n'
        '・業界平均値は中小企業庁データ等をもとにした参考値です。実際の業界平均と異なる場合があります。\n'
        '・差異の判定は「第3期実績 - 業界平均」で算出。利益率は高いほど良く、人件費率・固定比率は低いほど良い傾向です。\n'
        '・データが未入力の期は空欄になります。第3期（最新期）への入力が完了すると全指標が算出されます。'
    )
    note.font = mk_font(color='595959', size=8)
    note.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    note.border = mk_border()
    note.fill = mk_fill(C_LGRAY)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)

    ws.freeze_panes = 'C5'
    return ws


# ─────────────────────────────────────────────
# シート作成: 企業価値算定（修正前）
# ─────────────────────────────────────────────
def create_valuation_before_sheet(wb):
    ws_name = '企業価値算定（修正前）'
    if ws_name in wb.sheetnames:
        del wb[ws_name]
    ws = wb.create_sheet(ws_name)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    # ── タイトル ──
    title_row(ws, 1, '企業価値算定（修正前）　※帳簿値そのまま使用', col_to=6)

    ws.row_dimensions[2].height = 16
    ws.cell(2, 1).value = '単位：千円　　対象期：最新期（第3期）'
    ws.cell(2, 1).font = mk_font(color='595959', size=8)
    ws.cell(2, 1).alignment = mk_align(h='right')
    ws.merge_cells('A2:F2')

    # ────────────────────────────
    # セクション1: 基礎財務データ
    # ────────────────────────────
    R = 3
    section_row(ws, R, '【基礎財務データ（3期比較表より自動取得）】', col_to=6)

    base_items = [
        # (ラベル, 数式, 備考)
        ('売上高',        pl('売上高'),          '損益計算書より'),
        ('売上総利益',    pl('売上総利益'),       '損益計算書より'),
        ('営業利益',      pl('営業利益'),         '損益計算書より'),
        ('経常利益',      pl('経常利益'),         '損益計算書より'),
        ('当期純利益',    pl('当期純利益'),        '損益計算書より'),
        ('減価償却費',    sga('減価償却費'),       '販管費内の減価償却費'),
        ('純資産（簿価）',bs('純資産合計'),        '貸借対照表より'),
        ('資産合計',      bs('資産合計'),          '貸借対照表より'),
        ('負債合計',      bs('負債合計'),          '貸借対照表より'),
        ('流動負債合計',  bs('流動負債合計'),      '貸借対照表より'),
        ('固定負債合計',  bs('固定負債合計'),      '貸借対照表より'),
    ]

    # 行番号を記録（後で参照）
    data_rows = {}
    for i, (label, formula, note) in enumerate(base_items):
        row = R + 1 + i
        ws.row_dimensions[row].height = 17
        label_cell(ws.cell(row, 1), label, bold=True, fill=C_LLBLUE)
        c = ws.cell(row, 2)
        c.value = formula
        c.number_format = '#,##0'
        c.alignment = mk_align(h='right')
        c.border = mk_border()
        c.font = mk_font(bold=True)
        c.fill = mk_fill(C_LLBLUE)
        label_cell(ws.cell(row, 3), note, fill=C_LGRAY)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        data_rows[label] = row

    # EBITDA計算
    R_ebitda = R + 1 + len(base_items) + 1
    section_row(ws, R_ebitda - 1, '', col_to=6)  # セパレータ
    ws.row_dimensions[R_ebitda - 1].height = 6

    ws.row_dimensions[R_ebitda].height = 17
    label_cell(ws.cell(R_ebitda, 1), 'EBITDA（営業利益 + 減価償却費）', bold=True, fill=C_GOLD)
    c = ws.cell(R_ebitda, 2)
    op_row = data_rows['営業利益']
    dep_row = data_rows['減価償却費']
    c.value = f'=B{op_row}+B{dep_row}'
    c.number_format = '#,##0'
    c.alignment = mk_align(h='right')
    c.border = mk_border()
    c.font = mk_font(bold=True)
    c.fill = mk_fill(C_GOLD)
    label_cell(ws.cell(R_ebitda, 3), '※売上原価内の減価償却費は別途確認', fill=C_LGRAY)
    ws.merge_cells(start_row=R_ebitda, start_column=3, end_row=R_ebitda, end_column=6)
    ebitda_row = R_ebitda

    ws.row_dimensions[R_ebitda + 1].height = 17
    label_cell(ws.cell(R_ebitda + 1, 1), '有利子負債合計', bold=True, fill=C_LLBLUE)
    c = ws.cell(R_ebitda + 1, 2)
    c.value = f'=B{data_rows["固定負債合計"]}+B{data_rows["流動負債合計"]}'
    c.number_format = '#,##0'
    c.alignment = mk_align(h='right')
    c.border = mk_border()
    c.font = mk_font(bold=True)
    c.fill = mk_fill(C_LLBLUE)
    label_cell(ws.cell(R_ebitda + 1, 3), '※実際は有利子負債のみで計算（借入金等）', fill=C_LGRAY)
    ws.merge_cells(start_row=R_ebitda + 1, start_column=3, end_row=R_ebitda + 1, end_column=6)
    debt_row = R_ebitda + 1

    # 実効税率入力
    ws.row_dimensions[R_ebitda + 2].height = 17
    label_cell(ws.cell(R_ebitda + 2, 1), '実効税率（入力）', bold=True, fill=C_INPUT)
    c = ws.cell(R_ebitda + 2, 2)
    c.value = 0.30
    c.number_format = '0.0%'
    c.alignment = mk_align(h='right')
    c.border = mk_border()
    c.font = mk_font(bold=True, color='1F497D')
    c.fill = mk_fill(C_INPUT)
    label_cell(ws.cell(R_ebitda + 2, 3), '黄色セルは変更可能（デフォルト30%）', fill=C_LGRAY)
    ws.merge_cells(start_row=R_ebitda + 2, start_column=3, end_row=R_ebitda + 2, end_column=6)
    tax_row = R_ebitda + 2

    # ────────────────────────────
    # セクション2: 評価マルチプル
    # ────────────────────────────
    R_mult = tax_row + 2
    section_row(ws, R_mult, '【評価マルチプル（変更可能）】', col_to=6)

    mult_header_row = R_mult + 1
    ws.row_dimensions[mult_header_row].height = 20
    for ci, h in enumerate(['評価手法', 'マルチプル', '根拠・出典', '', '', ''], 1):
        hdr_cell(ws.cell(mult_header_row, ci), h)

    mult_items = [
        ('年買法　係数（低）',  2.0, '2年分の利益（保守的）'),
        ('年買法　係数（高）',  4.0, '4年分の利益（楽観的）'),
        ('EV/EBITDA倍率',      12.17, 'IT業界中央値（yfinance参照）'),
        ('EV/売上高倍率',       2.16, 'IT業界中央値（yfinance参照）'),
        ('PBR倍率',             2.87, 'IT業界中央値（yfinance参照）'),
        ('PER倍率',            23.42, 'IT業界中央値（yfinance参照）'),
    ]

    mult_rows = {}
    for i, (name, val, note) in enumerate(mult_items):
        row = mult_header_row + 1 + i
        ws.row_dimensions[row].height = 17
        label_cell(ws.cell(row, 1), name, fill=C_WHITE)
        c = ws.cell(row, 2)
        c.value = val
        c.number_format = '0.00'
        c.alignment = mk_align(h='right')
        c.border = mk_border()
        c.font = mk_font(color='1F497D', bold=True)
        c.fill = mk_fill(C_INPUT)
        label_cell(ws.cell(row, 3), note, fill=C_LGRAY)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        mult_rows[name] = row

    # ────────────────────────────
    # セクション3: 各手法による算定
    # ────────────────────────────
    R_val = mult_rows[mult_items[-1][0]] + 2
    section_row(ws, R_val, '【各評価手法による企業価値算定（千円）】', col_to=6)

    # ヘッダー
    val_hdr = R_val + 1
    ws.row_dimensions[val_hdr].height = 24
    val_headers = ['評価手法', '算出根拠', '事業価値(EV)', '株主価値(EQ)', '備考', '']
    for ci, h in enumerate(val_headers, 1):
        hdr_cell(ws.cell(val_hdr, ci), h)

    def val_row_write(row, name, ev_formula, note, dept_row_ref=None):
        ws.row_dimensions[row].height = 18
        label_cell(ws.cell(row, 1), name, bold=True)
        label_cell(ws.cell(row, 2), '', fill=C_LGRAY)

        c_ev = ws.cell(row, 3)
        c_ev.value = ev_formula
        c_ev.number_format = '#,##0'
        c_ev.alignment = mk_align(h='right')
        c_ev.border = mk_border()
        c_ev.font = mk_font(bold=True)
        c_ev.fill = mk_fill(C_GOLD)

        c_eq = ws.cell(row, 4)
        debt_ref = f'B{debt_row}'
        c_eq.value = f'=IFERROR(C{row}-{debt_ref},"")'
        c_eq.number_format = '#,##0'
        c_eq.alignment = mk_align(h='right')
        c_eq.border = mk_border()
        c_eq.font = mk_font(bold=True)
        c_eq.fill = mk_fill(C_GOLD)

        label_cell(ws.cell(row, 5), note, fill=C_LGRAY)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

    def r(name):
        return f'B{mult_rows[name]}'

    # 各手法
    cur = val_hdr + 1

    # 1. 年買法（低）
    net_asset_row = data_rows['純資産（簿価）']
    keijo_row = data_rows['経常利益']
    val_row_write(cur, '①年倍法（保守的・2年）',
        f'=IFERROR(B{net_asset_row}+B{keijo_row}*{r("年買法　係数（低）")},"")',
        '修正純資産 + 経常利益×2年')
    cur += 1

    # 2. 年買法（高）
    val_row_write(cur, '②年倍法（楽観的・4年）',
        f'=IFERROR(B{net_asset_row}+B{keijo_row}*{r("年買法　係数（高）")},"")',
        '修正純資産 + 経常利益×4年')
    cur += 1

    # 3. EV/EBITDA
    val_row_write(cur, '③EV/EBITDA法',
        f'=IFERROR(B{ebitda_row}*{r("EV/EBITDA倍率")},"")',
        f'EBITDA × {mult_items[2][1]}倍')
    cur += 1

    # 4. EV/売上高
    sales_row = data_rows['売上高']
    val_row_write(cur, '④EV/売上高法',
        f'=IFERROR(B{sales_row}*{r("EV/売上高倍率")},"")',
        f'売上高 × {mult_items[3][1]}倍')
    cur += 1

    # 5. PBR法（≒株主価値）
    ws.row_dimensions[cur].height = 18
    label_cell(ws.cell(cur, 1), '⑤PBR法（株主価値ベース）', bold=True)
    label_cell(ws.cell(cur, 2), '', fill=C_LGRAY)
    c_eq2 = ws.cell(cur, 4)
    c_eq2.value = f'=IFERROR(B{net_asset_row}*{r("PBR倍率")},"")'
    c_eq2.number_format = '#,##0'
    c_eq2.alignment = mk_align(h='right')
    c_eq2.border = mk_border()
    c_eq2.font = mk_font(bold=True)
    c_eq2.fill = mk_fill(C_GOLD)
    c_ev2 = ws.cell(cur, 3)
    c_ev2.value = f'=IFERROR(D{cur}+B{debt_row},"")'
    c_ev2.number_format = '#,##0'
    c_ev2.alignment = mk_align(h='right')
    c_ev2.border = mk_border()
    c_ev2.font = mk_font(bold=True)
    c_ev2.fill = mk_fill(C_GOLD)
    label_cell(ws.cell(cur, 5), f'純資産 × {mult_items[4][1]}倍', fill=C_LGRAY)
    ws.merge_cells(start_row=cur, start_column=5, end_row=cur, end_column=6)
    pbr_row = cur
    cur += 1

    # 6. PER法
    net_profit_row = data_rows['当期純利益']
    val_row_write(cur, '⑥PER法',
        f'=IFERROR(IF(B{net_profit_row}>0,B{net_profit_row}*{r("PER倍率")}+B{debt_row},"赤字のため算定不可"),"")',
        f'当期純利益 × {mult_items[5][1]}倍 + 有利子負債')
    per_row = cur
    cur += 1

    # 7. 純資産法
    ws.row_dimensions[cur].height = 18
    label_cell(ws.cell(cur, 1), '⑦純資産法（簿価）', bold=True)
    label_cell(ws.cell(cur, 2), '', fill=C_LGRAY)
    c_ev7 = ws.cell(cur, 3)
    c_ev7.value = f'=IFERROR(B{data_rows["資産合計"]}-B{data_rows["負債合計"]},"")'
    c_ev7.number_format = '#,##0'
    c_ev7.alignment = mk_align(h='right')
    c_ev7.border = mk_border()
    c_ev7.font = mk_font(bold=True)
    c_ev7.fill = mk_fill(C_GOLD)
    c_eq7 = ws.cell(cur, 4)
    c_eq7.value = f'=IFERROR(B{net_asset_row},"")'
    c_eq7.number_format = '#,##0'
    c_eq7.alignment = mk_align(h='right')
    c_eq7.border = mk_border()
    c_eq7.font = mk_font(bold=True)
    c_eq7.fill = mk_fill(C_GOLD)
    label_cell(ws.cell(cur, 5), '資産合計 - 負債合計', fill=C_LGRAY)
    ws.merge_cells(start_row=cur, start_column=5, end_row=cur, end_column=6)
    na_row = cur
    cur += 1

    # ── サマリー ──
    cur += 1
    section_row(ws, cur, '【評価レンジまとめ（株主価値ベース、千円）】', col_to=6)
    sum_hdr = cur + 1
    ws.row_dimensions[sum_hdr].height = 22

    hdr_cell(ws.cell(sum_hdr, 1), '指標')
    hdr_cell(ws.cell(sum_hdr, 2), '金額（千円）')
    hdr_cell(ws.cell(sum_hdr, 3), '年買法（低）')
    hdr_cell(ws.cell(sum_hdr, 4), '年買法（高）')
    hdr_cell(ws.cell(sum_hdr, 5), 'EV/EBITDA')
    hdr_cell(ws.cell(sum_hdr, 6), 'EV/売上高')

    # 株主価値サマリー行
    sum_labels = ['①年買法（低）', '②年買法（高）', '③EV/EBITDA法',
                  '④EV/売上高法', '⑤PBR法', '⑦純資産法']

    # 参照する株主価値のある行
    eq_refs = [val_hdr+1, val_hdr+2, val_hdr+3, val_hdr+4, pbr_row, na_row]

    cur2 = sum_hdr + 1
    ws.row_dimensions[cur2].height = 17
    label_cell(ws.cell(cur2, 1), '株主価値（最小）', bold=True, fill=C_LLBLUE)
    min_formula = 'MINIFS('
    eq_vals = ','.join([f'D{r}' for r in eq_refs])
    min_parts = ','.join([f'ISNUMBER(D{r}),TRUE' for r in eq_refs])
    # Simplified min formula
    vals_str = ','.join([f'IFERROR(D{r},"")'for r in eq_refs])
    c = ws.cell(cur2, 2)
    c.value = f'=IFERROR(MIN(D{eq_refs[0]},D{eq_refs[1]},D{eq_refs[2]},D{eq_refs[3]},D{eq_refs[4]},D{eq_refs[5]}),"")'
    c.number_format = '#,##0'
    c.alignment = mk_align(h='right')
    c.border = mk_border()
    c.font = mk_font(bold=True)
    c.fill = mk_fill(C_LLBLUE)
    ws.merge_cells(start_row=cur2, start_column=3, end_row=cur2, end_column=6)
    label_cell(ws.cell(cur2, 3), '各手法の最小値', fill=C_LGRAY)

    cur3 = cur2 + 1
    ws.row_dimensions[cur3].height = 17
    label_cell(ws.cell(cur3, 1), '株主価値（最大）', bold=True, fill=C_LLBLUE)
    c = ws.cell(cur3, 2)
    c.value = f'=IFERROR(MAX(D{eq_refs[0]},D{eq_refs[1]},D{eq_refs[2]},D{eq_refs[3]},D{eq_refs[4]},D{eq_refs[5]}),"")'
    c.number_format = '#,##0'
    c.alignment = mk_align(h='right')
    c.border = mk_border()
    c.font = mk_font(bold=True)
    c.fill = mk_fill(C_LLBLUE)
    ws.merge_cells(start_row=cur3, start_column=3, end_row=cur3, end_column=6)
    label_cell(ws.cell(cur3, 3), '各手法の最大値', fill=C_LGRAY)

    cur4 = cur3 + 1
    ws.row_dimensions[cur4].height = 22
    label_cell(ws.cell(cur4, 1), '推定評価レンジ', bold=True, fill=C_GOLD)
    c = ws.cell(cur4, 2)
    c.value = f'=IFERROR(TEXT(B{cur2},"#,##0")&" 〜 "&TEXT(B{cur3},"#,##0")&" 千円","データ不足")'
    c.font = mk_font(bold=True, color=C_NAVY, size=10)
    c.fill = mk_fill(C_GOLD)
    c.alignment = mk_align(h='center')
    c.border = mk_border()
    ws.merge_cells(start_row=cur4, start_column=2, end_row=cur4, end_column=6)

    # 注記
    note_r = cur4 + 2
    ws.row_dimensions[note_r].height = 80
    note = ws.cell(note_r, 1)
    note.value = (
        '【注記】\n'
        '・本算定は帳簿値（修正前）を使用しています。役員報酬の正常化・一時的損益の除外等の調整は「修正後」シートで行ってください。\n'
        '・有利子負債は全負債で代替しています。実際は借入金・社債等の有利子負債のみで計算してください。\n'
        '・マルチプルはIT業界上場会社の中央値（yfinance参照）を初期値としています。業種・規模に合わせて変更してください。\n'
        '・PER法・EV/EBITDA法は赤字の場合参考値となりません。年買法・純資産法を主軸に評価することを推奨します。\n'
        '・株主価値 = 事業価値(EV) - 有利子負債。EV = 事業に帰属する価値の総計。'
    )
    note.font = mk_font(color='595959', size=8)
    note.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    note.border = mk_border()
    note.fill = mk_fill(C_LGRAY)
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=6)

    ws.freeze_panes = 'A4'

    # 各評価手法行の修正後参照用に行番号を返す
    return {
        'ws': ws,
        'data_rows': data_rows,
        'ebitda_row': ebitda_row,
        'debt_row': debt_row,
        'tax_row': tax_row,
        'mult_rows': mult_rows,
        'val_hdr': val_hdr,
        'eq_refs': eq_refs,
        'pbr_row': pbr_row,
        'na_row': na_row,
    }


# ─────────────────────────────────────────────
# シート作成: 企業価値算定（修正後）
# ─────────────────────────────────────────────
def create_valuation_after_sheet(wb, before_info):
    ws_name = '企業価値算定（修正後）'
    if ws_name in wb.sheetnames:
        del wb[ws_name]
    ws = wb.create_sheet(ws_name)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    title_row(ws, 1, '企業価値算定（修正後）　※財務数値を正常化して算定', col_to=6)
    ws.row_dimensions[2].height = 14
    ws.cell(2, 1).value = '単位：千円　　黄色セルに調整額を入力してください（プラス＝利益増加・資産増加、マイナス＝利益減少・資産減少）'
    ws.cell(2, 1).font = mk_font(color='595959', size=8)
    ws.cell(2, 1).alignment = mk_align(h='right')
    ws.merge_cells('A2:F2')

    # ────────────────────────────
    # セクション1: 修正項目入力
    # ────────────────────────────
    R = 3
    section_row(ws, R, '【修正項目入力（黄色セルに調整額を入力）】', col_to=6)

    adj_hdr = R + 1
    ws.row_dimensions[adj_hdr].height = 22
    for ci, h in enumerate(['修正項目', '修正内容・理由', '調整額（千円）', '対象', ''], 1):
        hdr_cell(ws.cell(adj_hdr, ci), h)
    ws.merge_cells(start_row=adj_hdr, start_column=4, end_row=adj_hdr, end_column=6)

    adj_items = [
        # (ラベル, 説明, 対象, デフォルト)
        ('①役員報酬の正常化（利益調整）',
         '現役員報酬 - 同業他社適正水準（例: 過大報酬分を除外し利益を増加）',
         '経常利益に加算', 0),
        ('②オーナー家族への支払調整（利益）',
         '過大な家族役員報酬・地代家賃等の正常化',
         '経常利益に加算', 0),
        ('③一時的費用の除外（利益）',
         '再発しない特別費用（訴訟・災害・事業整理等）の除外',
         '経常利益に加算', 0),
        ('④一時的収益の除外（利益）',
         '再発しない特別収益（資産売却・補助金等）の除外',
         '経常利益に減算', 0),
        ('⑤役員退職慰労引当等（利益）',
         '過去分の引当超過や未引当調整',
         '経常利益に加減算', 0),
        ('⑥その他利益調整',
         '上記以外の利益正常化調整',
         '経常利益に加算', 0),
        ('⑦非事業用資産の除外（純資産）',
         '役員個人的保険・遊休資産・投資有価証券等',
         '純資産から除外（減算）', 0),
        ('⑧含み損益の反映（純資産）',
         '不動産・有価証券の時価と簿価の差額',
         '純資産に加算', 0),
        ('⑨簿外負債の考慮（純資産）',
         '未払退職金・リース・保証債務等の簿外負債',
         '純資産から控除（減算）', 0),
        ('⑩その他純資産調整',
         '上記以外の純資産正常化調整',
         '純資産に加算', 0),
    ]

    adj_rows = {}
    for i, (label, desc, target, default) in enumerate(adj_items):
        row = adj_hdr + 1 + i
        ws.row_dimensions[row].height = 18
        label_cell(ws.cell(row, 1), label, bold=False, fill=C_WHITE)
        label_cell(ws.cell(row, 2), desc, fill=C_LGRAY)
        c = ws.cell(row, 3)
        c.value = default if default != 0 else None
        c.fill = mk_fill(C_INPUT)
        c.font = mk_font(color='1F497D', bold=True)
        c.alignment = mk_align(h='right')
        c.border = mk_border()
        c.number_format = '#,##0'
        label_cell(ws.cell(row, 4), target, fill=C_LGRAY)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        adj_rows[label] = row

    adj_list = list(adj_rows.values())
    profit_adj_rows = adj_list[:6]   # ①~⑥: 利益調整
    asset_adj_rows  = adj_list[6:]   # ⑦~⑩: 純資産調整

    # ────────────────────────────
    # セクション2: 修正後財務数値
    # ────────────────────────────
    R2 = adj_hdr + 1 + len(adj_items) + 1
    section_row(ws, R2, '【修正後財務数値（自動計算）】', col_to=6)

    # 参照先（修正前シート）の行番号
    bef = before_info
    bef_ws = '企業価値算定（修正前）'

    def bef_ref(row_num, col='B'):
        return f"=IFERROR('{bef_ws}'!{col}{row_num},0)"

    adj_data_items = [
        # (ラベル, 修正前参照, 調整額式)
        ('修正前　経常利益',
         bef_ref(bef['data_rows']['経常利益']), None),
        ('利益調整額合計（①〜⑥）',
         None,
         '=' + '+'.join([f'IFERROR(C{r},0)' for r in profit_adj_rows])),
        ('修正後　経常利益',
         None, None),  # 計算式
        ('修正前　純資産（簿価）',
         bef_ref(bef['data_rows']['純資産（簿価）']), None),
        ('純資産調整額合計（⑦〜⑩）',
         None,
         '=' + '+'.join([f'IFERROR(C{r},0)' for r in asset_adj_rows])),
        ('修正後　純資産',
         None, None),  # 計算式
        ('修正後　EBITDA',
         None, None),  # 計算式
        ('有利子負債合計',
         bef_ref(bef['debt_row']), None),
    ]

    adj_data_rows = {}
    cur = R2 + 1
    for label, base_formula, adj_formula in adj_data_items:
        ws.row_dimensions[cur].height = 17
        is_total = '合計' in label or '修正後' in label
        fill = C_GOLD if '修正後' in label else (C_LLBLUE if '合計' in label else C_WHITE)
        label_cell(ws.cell(cur, 1), label, bold=is_total, fill=fill)

        c = ws.cell(cur, 3)
        if base_formula:
            c.value = base_formula
        elif adj_formula:
            c.value = adj_formula
        c.number_format = '#,##0'
        c.alignment = mk_align(h='right')
        c.border = mk_border()
        c.font = mk_font(bold=is_total)
        c.fill = mk_fill(fill)

        # B列（ラベル補足）
        ws.cell(cur, 2).border = mk_border()
        ws.cell(cur, 2).fill = mk_fill(C_LGRAY)

        adj_data_rows[label] = cur
        cur += 1

    # 修正後経常利益 = 修正前 + 調整
    r_adj_keijo = adj_data_rows['修正後　経常利益']
    r_base_keijo = adj_data_rows['修正前　経常利益']
    r_kadj = adj_data_rows['利益調整額合計（①〜⑥）']
    ws.cell(r_adj_keijo, 3).value = f'=C{r_base_keijo}+C{r_kadj}'
    ws.cell(r_adj_keijo, 3).number_format = '#,##0'
    ws.cell(r_adj_keijo, 3).alignment = mk_align(h='right')
    ws.cell(r_adj_keijo, 3).border = mk_border()
    ws.cell(r_adj_keijo, 3).font = mk_font(bold=True)
    ws.cell(r_adj_keijo, 3).fill = mk_fill(C_GOLD)
    adj_data_rows['修正後　経常利益'] = r_adj_keijo

    # 修正後純資産 = 修正前 + 調整
    r_adj_na = adj_data_rows['修正後　純資産']
    r_base_na = adj_data_rows['修正前　純資産（簿価）']
    r_naadj = adj_data_rows['純資産調整額合計（⑦〜⑩）']
    ws.cell(r_adj_na, 3).value = f'=C{r_base_na}+C{r_naadj}'
    ws.cell(r_adj_na, 3).number_format = '#,##0'
    ws.cell(r_adj_na, 3).alignment = mk_align(h='right')
    ws.cell(r_adj_na, 3).border = mk_border()
    ws.cell(r_adj_na, 3).font = mk_font(bold=True)
    ws.cell(r_adj_na, 3).fill = mk_fill(C_GOLD)

    # 修正後EBITDA
    r_adj_ebitda = adj_data_rows['修正後　EBITDA']
    dep_ref = f"IFERROR('{bef_ws}'!B{bef['data_rows']['減価償却費']},0)"
    ws.cell(r_adj_ebitda, 3).value = f'=C{r_adj_keijo}+{dep_ref}'
    ws.cell(r_adj_ebitda, 3).number_format = '#,##0'
    ws.cell(r_adj_ebitda, 3).alignment = mk_align(h='right')
    ws.cell(r_adj_ebitda, 3).border = mk_border()
    ws.cell(r_adj_ebitda, 3).font = mk_font(bold=True)
    ws.cell(r_adj_ebitda, 3).fill = mk_fill(C_GOLD)

    # 有利子負債（参照）
    r_debt_adj = adj_data_rows['有利子負債合計']

    # ────────────────────────────
    # セクション3: 修正後評価
    # ────────────────────────────
    R3 = cur + 1
    section_row(ws, R3, '【修正後　各評価手法による企業価値算定（千円）】', col_to=6)

    val_hdr2 = R3 + 1
    ws.row_dimensions[val_hdr2].height = 24
    for ci, h in enumerate(['評価手法', '算出根拠', '事業価値(EV)', '株主価値(EQ)', '修正前比較', ''], 1):
        hdr_cell(ws.cell(val_hdr2, ci), h)

    # マルチプル参照（修正前シートから）
    def m_ref(name):
        return f"IFERROR('{bef_ws}'!B{bef['mult_rows'][name]},0)"

    sales_row_bef = bef['data_rows']['売上高']

    def val2_row(row, name, ev_formula, note, before_eq_row=None):
        ws.row_dimensions[row].height = 18
        label_cell(ws.cell(row, 1), name, bold=True)
        label_cell(ws.cell(row, 2), '', fill=C_LGRAY)

        c_ev = ws.cell(row, 3)
        c_ev.value = ev_formula
        c_ev.number_format = '#,##0'
        c_ev.alignment = mk_align(h='right')
        c_ev.border = mk_border()
        c_ev.font = mk_font(bold=True)
        c_ev.fill = mk_fill(C_GOLD)

        c_eq = ws.cell(row, 4)
        c_eq.value = f'=IFERROR(C{row}-C{r_debt_adj},"")'
        c_eq.number_format = '#,##0'
        c_eq.alignment = mk_align(h='right')
        c_eq.border = mk_border()
        c_eq.font = mk_font(bold=True)
        c_eq.fill = mk_fill(C_GOLD)

        # 修正前比較
        c_cmp = ws.cell(row, 5)
        if before_eq_row:
            c_cmp.value = f"=IFERROR(D{row}-'{bef_ws}'!D{before_eq_row},\"\")"
            c_cmp.number_format = '#,##0'
            c_cmp.alignment = mk_align(h='right')
            c_cmp.border = mk_border()
            c_cmp.font = mk_font()
            c_cmp.fill = mk_fill(C_LGRAY)
            # 条件付き書式（差分が正なら緑、負なら赤）
        else:
            label_cell(ws.cell(row, 5), '─', fill=C_LGRAY)
        label_cell(ws.cell(row, 6), note, fill=C_LGRAY)

    cur_v = val_hdr2 + 1
    bef_val_start = bef['val_hdr'] + 1
    before_eq_rows_list = bef['eq_refs']

    # ①年買法（低）
    val2_row(cur_v, '①年倍法（保守的・2年）',
        f'=IFERROR(C{r_adj_na}+C{r_adj_keijo}*{m_ref("年買法　係数（低）")},"")',
        '修正純資産 + 修正経常利益×2年',
        before_eq_rows_list[0])
    cur_v += 1

    # ②年買法（高）
    val2_row(cur_v, '②年倍法（楽観的・4年）',
        f'=IFERROR(C{r_adj_na}+C{r_adj_keijo}*{m_ref("年買法　係数（高）")},"")',
        '修正純資産 + 修正経常利益×4年',
        before_eq_rows_list[1])
    cur_v += 1

    # ③EV/EBITDA
    val2_row(cur_v, '③EV/EBITDA法',
        f'=IFERROR(C{r_adj_ebitda}*{m_ref("EV/EBITDA倍率")},"")',
        '修正EBITDA × 倍率',
        before_eq_rows_list[2])
    cur_v += 1

    # ④EV/売上高
    val2_row(cur_v, '④EV/売上高法',
        f"=IFERROR('{bef_ws}'!B{sales_row_bef}*{m_ref('EV/売上高倍率')},\"\")",
        '売上高 × 倍率（売上は修正不要）',
        before_eq_rows_list[3])
    cur_v += 1

    # ⑤PBR法
    ws.row_dimensions[cur_v].height = 18
    label_cell(ws.cell(cur_v, 1), '⑤PBR法（株主価値ベース）', bold=True)
    label_cell(ws.cell(cur_v, 2), '', fill=C_LGRAY)
    c_eq5 = ws.cell(cur_v, 4)
    c_eq5.value = f'=IFERROR(C{r_adj_na}*{m_ref("PBR倍率")},"")'
    c_eq5.number_format = '#,##0'
    c_eq5.alignment = mk_align(h='right')
    c_eq5.border = mk_border()
    c_eq5.font = mk_font(bold=True)
    c_eq5.fill = mk_fill(C_GOLD)
    c_ev5 = ws.cell(cur_v, 3)
    c_ev5.value = f'=IFERROR(D{cur_v}+C{r_debt_adj},"")'
    c_ev5.number_format = '#,##0'
    c_ev5.alignment = mk_align(h='right')
    c_ev5.border = mk_border()
    c_ev5.font = mk_font(bold=True)
    c_ev5.fill = mk_fill(C_GOLD)
    c_cmp5 = ws.cell(cur_v, 5)
    c_cmp5.value = f"=IFERROR(D{cur_v}-'{bef_ws}'!D{bef['pbr_row']},\"\")"
    c_cmp5.number_format = '#,##0'
    c_cmp5.alignment = mk_align(h='right')
    c_cmp5.border = mk_border()
    c_cmp5.font = mk_font()
    c_cmp5.fill = mk_fill(C_LGRAY)
    label_cell(ws.cell(cur_v, 6), '修正純資産 × PBR', fill=C_LGRAY)
    pbr2_row = cur_v
    cur_v += 1

    # ⑦純資産法
    ws.row_dimensions[cur_v].height = 18
    label_cell(ws.cell(cur_v, 1), '⑦純資産法（修正後）', bold=True)
    label_cell(ws.cell(cur_v, 2), '', fill=C_LGRAY)
    c_ev7 = ws.cell(cur_v, 3)
    c_ev7.value = f'=IFERROR(C{r_adj_na}+C{r_debt_adj},"")'
    c_ev7.number_format = '#,##0'
    c_ev7.alignment = mk_align(h='right')
    c_ev7.border = mk_border()
    c_ev7.font = mk_font(bold=True)
    c_ev7.fill = mk_fill(C_GOLD)
    c_eq7 = ws.cell(cur_v, 4)
    c_eq7.value = f'=IFERROR(C{r_adj_na},"")'
    c_eq7.number_format = '#,##0'
    c_eq7.alignment = mk_align(h='right')
    c_eq7.border = mk_border()
    c_eq7.font = mk_font(bold=True)
    c_eq7.fill = mk_fill(C_GOLD)
    c_cmp7 = ws.cell(cur_v, 5)
    c_cmp7.value = f"=IFERROR(D{cur_v}-'{bef_ws}'!D{bef['na_row']},\"\")"
    c_cmp7.number_format = '#,##0'
    c_cmp7.alignment = mk_align(h='right')
    c_cmp7.border = mk_border()
    c_cmp7.font = mk_font()
    c_cmp7.fill = mk_fill(C_LGRAY)
    label_cell(ws.cell(cur_v, 6), '修正純資産', fill=C_LGRAY)
    na2_row = cur_v
    cur_v += 1

    # ── サマリー ──
    cur_v += 1
    section_row(ws, cur_v, '【修正後　評価レンジまとめ（株主価値ベース、千円）】', col_to=6)
    sum2_hdr = cur_v + 1
    ws.row_dimensions[sum2_hdr].height = 22
    for ci, h in enumerate(['指標', '修正後（千円）', '修正前（千円）', '差異（千円）', '', ''], 1):
        hdr_cell(ws.cell(sum2_hdr, ci), h)

    eq2_refs = [val_hdr2+1, val_hdr2+2, val_hdr2+3, val_hdr2+4, pbr2_row, na2_row]

    cur_s = sum2_hdr + 1
    ws.row_dimensions[cur_s].height = 17
    label_cell(ws.cell(cur_s, 1), '株主価値（最小）', bold=True, fill=C_LLBLUE)
    min_v = f'=IFERROR(MIN({",".join([f"D{r}" for r in eq2_refs])}),"")'
    c = ws.cell(cur_s, 2)
    c.value = min_v
    c.number_format = '#,##0'
    c.alignment = mk_align(h='right')
    c.border = mk_border()
    c.font = mk_font(bold=True)
    c.fill = mk_fill(C_LLBLUE)
    bef_min_rows = bef['eq_refs']
    c3 = ws.cell(cur_s, 3)
    _bef_min_refs = ','.join(["'" + bef_ws + "'!D" + str(r) for r in bef_min_rows])
    c3.value = f'=IFERROR(MIN({_bef_min_refs}),"")'
    c3.number_format = '#,##0'
    c3.alignment = mk_align(h='right')
    c3.border = mk_border()
    c3.font = mk_font()
    c3.fill = mk_fill(C_LGRAY)
    c4 = ws.cell(cur_s, 4)
    c4.value = f'=IFERROR(B{cur_s}-C{cur_s},"")'
    c4.number_format = '#,##0'
    c4.alignment = mk_align(h='right')
    c4.border = mk_border()
    c4.font = mk_font()
    c4.fill = mk_fill(C_LGRAY)
    ws.merge_cells(start_row=cur_s, start_column=5, end_row=cur_s, end_column=6)
    label_cell(ws.cell(cur_s, 5), '各手法の最小値', fill=C_LGRAY)

    cur_s2 = cur_s + 1
    ws.row_dimensions[cur_s2].height = 17
    label_cell(ws.cell(cur_s2, 1), '株主価値（最大）', bold=True, fill=C_LLBLUE)
    max_v = f'=IFERROR(MAX({",".join([f"D{r}" for r in eq2_refs])}),"")'
    c = ws.cell(cur_s2, 2)
    c.value = max_v
    c.number_format = '#,##0'
    c.alignment = mk_align(h='right')
    c.border = mk_border()
    c.font = mk_font(bold=True)
    c.fill = mk_fill(C_LLBLUE)
    c3 = ws.cell(cur_s2, 3)
    _bef_max_refs = ','.join(["'" + bef_ws + "'!D" + str(r) for r in bef_min_rows])
    c3.value = f'=IFERROR(MAX({_bef_max_refs}),"")'
    c3.number_format = '#,##0'
    c3.alignment = mk_align(h='right')
    c3.border = mk_border()
    c3.font = mk_font()
    c3.fill = mk_fill(C_LGRAY)
    c4 = ws.cell(cur_s2, 4)
    c4.value = f'=IFERROR(B{cur_s2}-C{cur_s2},"")'
    c4.number_format = '#,##0'
    c4.alignment = mk_align(h='right')
    c4.border = mk_border()
    c4.font = mk_font()
    c4.fill = mk_fill(C_LGRAY)
    ws.merge_cells(start_row=cur_s2, start_column=5, end_row=cur_s2, end_column=6)
    label_cell(ws.cell(cur_s2, 5), '各手法の最大値', fill=C_LGRAY)

    cur_s3 = cur_s2 + 1
    ws.row_dimensions[cur_s3].height = 24
    label_cell(ws.cell(cur_s3, 1), '推定評価レンジ（修正後）', bold=True, fill=C_GOLD)
    c = ws.cell(cur_s3, 2)
    c.value = f'=IFERROR(TEXT(B{cur_s},"#,##0")&" 〜 "&TEXT(B{cur_s2},"#,##0")&" 千円","データ不足")'
    c.font = mk_font(bold=True, color=C_NAVY, size=10)
    c.fill = mk_fill(C_GOLD)
    c.alignment = mk_align(h='center')
    c.border = mk_border()
    ws.merge_cells(start_row=cur_s3, start_column=2, end_row=cur_s3, end_column=6)

    # E列（差異）の条件付き書式
    last_val_row = cur_s2
    ws.conditional_formatting.add(f'E{val_hdr2+1}:E{last_val_row}',
        FormulaRule(formula=[f'E{val_hdr2+1}>0'],
                    fill=mk_fill('C6EFCE'), font=mk_font(color=C_DGRN)))
    ws.conditional_formatting.add(f'E{val_hdr2+1}:E{last_val_row}',
        FormulaRule(formula=[f'E{val_hdr2+1}<0'],
                    fill=mk_fill('FFC7CE'), font=mk_font(color=C_DRED)))

    # 注記
    note_r = cur_s3 + 2
    ws.row_dimensions[note_r].height = 80
    note = ws.cell(note_r, 1)
    note.value = (
        '【修正項目の考え方】\n'
        '①役員報酬の正常化: 現役員報酬が同業他社水準より過大な場合、差額分を利益に戻す（利益=増加→プラス入力）\n'
        '②家族への過大支払: オーナー家族への過大な役員報酬・地代家賃等を正常化（プラス入力で利益増加）\n'
        '③④一時的損益: 事業承継後に繰り返さない収益・費用を除外（費用除外→プラス、収益除外→マイナス）\n'
        '⑦非事業用資産: 事業に不要な資産（役員保険積立金等）を除外（マイナス入力で純資産減少）\n'
        '⑧含み損益: 不動産や有価証券の時価と簿価の差（含み益→プラス、含み損→マイナス）\n'
        '⑨簿外負債: 未払退職金・保証債務等（マイナス入力で純資産減少）'
    )
    note.font = mk_font(color='595959', size=8)
    note.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    note.border = mk_border()
    note.fill = mk_fill(C_LGRAY)
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=6)

    ws.freeze_panes = 'A4'
    return ws


# ─────────────────────────────────────────────
# メイン処理
# ─────────────────────────────────────────────
def main():
    input_file  = '3期比較表.xlsx'
    output_file = '決算分析ツール.xlsx'

    if not os.path.exists(input_file):
        print(f'エラー: {input_file} が見つかりません')
        sys.exit(1)

    print(f'■ {input_file} を読み込み中...')
    shutil.copy2(input_file, output_file)
    wb = openpyxl.load_workbook(output_file)

    # 既存の「企業価値算定」シートを削除（置き換え）
    for name in ['企業価値算定', '業界平均比較・アラート',
                 '企業価値算定（修正前）', '企業価値算定（修正後）', '業界データ']:
        if name in wb.sheetnames:
            del wb[name]

    print('■ 業界データシートを作成中...')
    create_industry_data_sheet(wb)

    print('■ 業界平均比較・アラートシートを作成中...')
    create_industry_comparison_sheet(wb)

    print('■ 企業価値算定（修正前）シートを作成中...')
    before_info = create_valuation_before_sheet(wb)

    print('■ 企業価値算定（修正後）シートを作成中...')
    create_valuation_after_sheet(wb, before_info)

    # シート順序を整理（openpyxlの内部リストを直接ソート）
    desired_order = [
        '貸借対照表', '損益計算書', '販売費及び一般管理費',
        '業界平均比較・アラート', '企業価値算定（修正前）', '企業価値算定（修正後）',
        '業界データ',
    ]
    def sort_key(ws):
        try:
            return desired_order.index(ws.title)
        except ValueError:
            return len(desired_order)
    wb._sheets.sort(key=sort_key)

    wb.save(output_file)
    print(f'\n[完了] {output_file} を開いてください')
    print()
    print('【使い方】')
    print('  1. 貸借対照表・損益計算書・販管費シートに決算数値を入力')
    print('  2. 業界平均比較・アラートシートで業種を選択 → 自動でアラート表示')
    print('  3. 企業価値算定（修正前）で帳簿値ベースの評価を確認')
    print('  4. 企業価値算定（修正後）で調整額を入力 → 修正後評価を確認')


if __name__ == '__main__':
    main()
